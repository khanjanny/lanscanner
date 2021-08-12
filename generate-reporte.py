#!/usr/bin/python3
# Ejecutar en la carpeta raiz (en el nivel de las carpetas EXTERNO/INTERNO )
import os
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.chart.marker import DataPoint
from openpyxl.chart.label import DataLabelList
from lxml import etree
import subprocess
import argparse
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.chart import PieChart3D, PieChart, ProjectedPieChart, BarChart, Reference
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.chart.marker import DataPoint
from openpyxl.chart.label import DataLabelList
import sqlite3
import pprint



parser = argparse.ArgumentParser()
# Add long and short argument
parser.add_argument("--domain", "-d", help="set domain")
args = parser.parse_args()
matrizRecomendaciones = []

# Definir los archivos para los informes
empresas = ["Radical","BISIT"]
wbEjecutivo = openpyxl.Workbook() # el informe ejecutivo es igual para todos

# Fuentes
MuyAlto = Font(name='Arial', size=12, bold=True, color='FC611C')
Alto = Font(name='Arial', size=12, bold=True, color='FFC000')
Moderado = Font(name='Arial', size=12, bold=True, color='FFFF00')
Bajo = Font(name='Arial', size=12, bold=True, color='95B3D7')
MuyBajo = Font(name='Arial', size=12, bold=True, color='92D050')

italic24Font = Font(size=24, italic=True, bold=True)
Arial10 = Font(name='Arial', size=10)
Arial11Bold = Font(name='Arial', size=11, bold=True)
Arial12BoldWhite = Font(name='Arial', size=12, bold=True, color='FFFFFF')
Arial10Bold = Font(name='Arial', size=10, bold=True)
Calibri10 = Font(name='Calibri', size=10)
Calibri10Bold = Font(name='Calibri', size=10, bold=True)
Calibri12Bold = Font(name='Calibri', size=12, bold=True)
Calibri10BoldWhite = Font(name='Calibri', size=10, bold=True, color='FFFFFF')

Impact22 = Font(name='Impact', size=22)
Impact36 = Font(name='Impact', size=36)

# Borde
thin_border = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))

# Rellenos
skyBlueFill = PatternFill(start_color='bdd7ee', end_color='bdd7ee', fill_type='solid')
greyFill = PatternFill(start_color='d9d9d9', end_color='d9d9d9', fill_type='solid')
verdeFill = PatternFill(start_color='31849B', end_color='31849B', fill_type='solid')

#### Radical #####
criticoFill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
altoFill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
medioFill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
bajoFill = PatternFill(start_color='00B050', end_color='00B050', fill_type='solid')
brownFill = PatternFill(start_color='EEECE1', end_color='EEECE1', fill_type='solid')

#### BISIT #####
criticoFill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
AltoFill2 = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
ModeradoFill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
bajoFill2 = PatternFill(start_color='92D050', end_color='92D050', fill_type='solid')
MuyBajoFill = PatternFill(start_color='95B3D7', end_color='95B3D7', fill_type='solid')
negroFill = PatternFill(start_color='17150D', end_color='17150D', fill_type='solid')

############################################################

# Busca el codigo CVE y la descripcion de una vulnerabilidad
def buscarDescripcion(nombreVul):
    codVul = "N/A"
    for i in range(total_vulnerabilidades):
        current_cod = root[i].find("cod").text
        descripcion = root[i].find("nombre").text
        if nombreVul in current_cod:
            codVul = root[i].find("codVul").text
            break
    if "webarchivos" == nombreVul: descripcion = "Navegación forzada para descubrir archivos comunes"
    if "passwordAdivinado1" == nombreVul: descripcion = "Contraseña fácil de adivinar"
    if "passwordAdivinado2" == nombreVul: descripcion = "Contraseña fácil de adivinar"
    if "webdirectorios" == nombreVul: descripcion = "Navegación forzada para descubrir archivos directorios"
    if "googlehacking0" == nombreVul: descripcion = "Google dork: site:DOMINIO inurl:add"
    if "googlehacking1" == nombreVul: descripcion = "Google dork: site:DOMINIO inurl:edit"
    if "googlehacking2" == nombreVul: descripcion = "Google dork: site:DOMINIO intitle:index.of"
    if "googlehacking3" == nombreVul: descripcion = "Google dork: site:DOMINIO filetype:sql"
    if "googlehacking4" == nombreVul: descripcion = 'Google dork: site:DOMINIO "access denied for user"'
    if "googlehacking5" == nombreVul: descripcion = 'Google dork: site:DOMINIO intitle:"curriculum vitae"'
    if "googlehacking6" == nombreVul: descripcion = "Google dork: site:DOMINIO passwords|contrasenas|login|contrasena filetype:txt"
    if "googlehacking11" == nombreVul: descripcion = "Google dork: site:trello.com passwords|contrasenas|login|contrasena intext:DOMINIO"
    if "googlehacking13" == nombreVul: descripcion = 'Google dork: site:DOMINIO "Undefined index"'
    if "googlehacking14" == nombreVul: descripcion = 'Google dork: site:DOMINIO inurl:storage'
    return codVul+";"+descripcion



root = etree.parse("/usr/share/lanscanner/vulnerabilidades.xml").getroot()
total_vulnerabilidades = len(root.getchildren())
print (f"total_vulnerabilidades en xml {total_vulnerabilidades}")

# Resultados Externos
result = subprocess.run(['find', 'EXTERNO','-iname', '.resultados.db'], stdout=subprocess.PIPE)
resultados_externos = result.stdout.decode("utf-8")
#print(f"resultados_externos {resultados_externos}")
resultados_externos_list = resultados_externos.split('\n')
resultados_externos_list.pop() # eliminar ultimo elemento (vacio)

# Resultados Internos
result = subprocess.run(['find', 'INTERNO','-iname', '.resultados.db'], stdout=subprocess.PIPE)
resultados_internos = result.stdout.decode("utf-8")
#print(f"resultados_externos {resultados_internos}")
resultados_internos_list = resultados_internos.split('\n')
resultados_internos_list.pop() # eliminar ultimo elemento (vacio)

todos_resultados = resultados_internos_list + resultados_externos_list

#Para el Informe ejecutivo
# TODAS LAS PRUEBAS
total_host_analizados = 0
total_host_vulnerabilidades = 0
total_host_uniq_vulnerabilidades = 0

######## #Vulnerabilidades por riesgo ########
total_vul_criticas = 0
total_vul_altas = 0
total_vul_medias = 0
total_vul_bajas = 0
######

######## total host afectados  ########
total_servicios_vuln_criticas = 0
total_servicios_vuln_altas = 0
total_servicios_vuln_medias = 0
total_servicios_vuln_bajas = 0
######
#####################################

###### Vulnerabilidades por vector
total_vuln_externas = 0
total_vuln_internas = 0
#######

total_vul_criticas_EXTERNO = 0
total_vul_altas_EXTERNO = 0
total_vul_medias_EXTERNO  = 0
total_vul_bajas_EXTERNO = 0

total_vul_criticas_INTERNO = 0
total_vul_altas_INTERNO= 0
total_vul_medias_INTERNO  = 0
total_vul_bajas_INTERNO = 0


###### Vulnerabilidades por activos ####
aplicacionWeb = 0
servidores = 0
baseDatos = 0
estacionesTrabajo = 0
telefoniaIP = 0
sistemaVigilancia = 0
dispositivosRed = 0
personal = 0
otros = 0  # Impresoras, lectores de huella
#############

#### Vulnerabilidades por categoria de vulnerabilidad ####
passwordDebil = 0
faltaParches = 0
errorConfiguracion = 0
####################
totalPruebas = 0
for resultados_db in todos_resultados:

    host_analizados = 0

    print(f"resultados_db {resultados_db}")
    ruta = resultados_db
    ruta = ruta.replace(".resultados.db", "")
    vectorInforme = resultados_db.split('/')[0]
    print(f"ruta {ruta}")
    print(f"vectorInforme {vectorInforme}")


    conn = sqlite3.connect(resultados_db)
    c = conn.cursor()

    #vul_externas = c.execute('select COUNT (DISTINCT TIPO) from VULNERABILIDADES').fetchone()[0]
    stream = os.popen('wc -l ' + ruta + '/.datos/total-host-vivos.txt | cut -d " " -f1')
    host_analizados = int(stream.read())
    total_host_analizados = total_host_analizados + host_analizados
    print(f"total_host_analizados {total_host_analizados}")

    # Host unicos con alguna vulnerabilidad
    host_uniq_vulnerabilidades = c.execute('SELECT COUNT (DISTINCT IP) FROM VULNERABILIDADES').fetchone()[0]
    print(f"host_uniq_vulnerabilidades {host_uniq_vulnerabilidades}")

    #mmmm
    #host_vulnerabilidades = c.execute('SELECT COUNT (IP) FROM VULNERABILIDADES;').fetchone()[0]
    #print(f"host_vulnerabilidades {host_vulnerabilidades}")

    ########################### Vulnerabilidades por criticidad #################

    # Vulnerabilidades criticas unicas
    vul_criticas = c.execute("select COUNT (DISTINCT TIPO) from VULNERABILIDADES where tipo ='ms08067'  or tipo ='passTomcat' or tipo ='zimbraXXE' or tipo ='doublepulsar' or tipo ='webshell' or tipo ='backdoorFabrica' or tipo ='ransomware' or tipo ='passwordSFI' or tipo ='BlueKeep' or tipo ='JoomlaJCKeditor' or tipo ='zerologon'").fetchone()[0]
    total_vul_criticas = total_vul_criticas + vul_criticas
    if "INTERNO" in vectorInforme: total_vul_criticas_INTERNO = total_vul_criticas_INTERNO + vul_criticas
    if "EXTERNO" in vectorInforme: total_vul_criticas_EXTERNO = total_vul_criticas_EXTERNO + vul_criticas
    print(f"total_vul_criticas {total_vul_criticas}")


    # Vulnerabilidades altas unicas
    vul_altos = c.execute("select COUNT (DISTINCT TIPO) from VULNERABILIDADES where tipo ='ms17010' or tipo ='archivosPeligrosos' or tipo ='mailPass' or tipo ='passwordDefecto' or tipo ='compartidoNFS' or tipo ='compartidoSMB' or tipo ='passwordHost' or tipo ='logeoRemoto' or tipo ='heartbleed' or tipo ='passwordAdivinado' or tipo ='passwordMikroTik' or tipo ='VNCnopass' or tipo ='VNCbypass' or tipo ='vulnDahua' or tipo ='openrelay' or tipo ='perdidaAutenticacion' or tipo ='spoof' or tipo ='slowloris' or tipo ='wordpressPass' or tipo ='conficker' or tipo ='anonymousIPMI' or tipo ='noSQLDatabases' or tipo ='winboxVuln' or tipo ='rmiVuln' or tipo ='SSHBypass' or tipo ='intelVuln' or tipo ='backupWeb' or tipo ='apacheStruts' or tipo ='webdavVulnerable' or tipo ='IISwebdavVulnerable' or tipo ='shellshock' or tipo ='ciscoASAVuln' or tipo ='SambaCry' or tipo ='misfortune' or tipo ='jbossVuln' or tipo ='passwordBD' or tipo ='wordpressDesactualizado'  or tipo ='llmnr' or tipo ='poisoning' or tipo ='cipherZeroIPMI' or tipo ='vlanHop' or tipo ='owaVul' or tipo ='phishing' or tipo ='hashRoto' or tipo ='pluginDesactualizado' or tipo ='cmsDesactualizado'").fetchone()[0]
    total_vul_altas = total_vul_altas + vul_altos
    if "INTERNO" in vectorInforme: total_vul_altas_INTERNO = total_vul_altas_INTERNO + vul_altos
    if "EXTERNO" in vectorInforme: total_vul_altas_EXTERNO = total_vul_altas_EXTERNO + vul_altos
    print(f"total_vul_altas {total_vul_altas}")

    # Vulnerabilidades medias unicas
    vul_medias = c.execute("select COUNT (DISTINCT TIPO) from VULNERABILIDADES where tipo ='VPNhandshake' or tipo ='openstreaming'  or tipo ='directorioLDAP' or tipo ='enum4linux' or tipo ='transferenciaDNS' or tipo ='listadoDirectorio' or tipo ='enumeracionUsuarios' or tipo ='googlehacking' or tipo ='anonymous' or tipo ='ACL' or tipo ='openresolver' or tipo ='ms12020' or tipo ='wpusers' or tipo ='exposicionUsuarios' or tipo ='HTTPsys' or tipo ='upnpAbierto' or tipo ='registroHabilitado' or tipo ='stp' or tipo ='noTLS' or tipo ='wordpressPingbacks' or tipo ='vulTLS'").fetchone()[0]
    total_vul_medias = total_vul_medias + vul_medias
    if "INTERNO" in vectorInforme: total_vul_medias_INTERNO = total_vul_medias_INTERNO + vul_medias
    if "EXTERNO" in vectorInforme: total_vul_medias_EXTERNO = total_vul_medias_EXTERNO + vul_medias
    print(f"total_vul_medias {total_vul_medias}")

    # Vulnerabilidades bajas unicas
    vul_bajas = c.execute("select COUNT (DISTINCT TIPO) from VULNERABILIDADES where tipo ='passwordDahua' or tipo ='divulgacionInformacion'  or tipo ='snmpCommunity'  or tipo ='listadoDirectorio' or tipo ='enumeracionUsuarios' or tipo ='googlehacking' or tipo ='anonymous' or tipo ='erroresWeb' or tipo ='ACL' or tipo ='archivosDefecto' or tipo ='openresolver' or tipo ='listadoDirectorios' or tipo ='debugHabilitado' or tipo ='CVE15473' or tipo ='IPinterna' or tipo ='ftpAnonymous' or tipo ='contenidoNoRelacionado' or tipo ='confTLS' or tipo ='captcha' or tipo ='contenidoIndebido' ").fetchone()[0]
    total_vul_bajas = total_vul_bajas + vul_bajas
    if "INTERNO" in vectorInforme: total_vul_bajas_INTERNO = total_vul_bajas_INTERNO + vul_bajas
    if "EXTERNO" in vectorInforme: total_vul_bajas_EXTERNO = total_vul_bajas_EXTERNO + vul_bajas
    print(f"total_vul_bajas {total_vul_bajas}")
    ###################################################

    ############################ Servicios afectados por vulnerabilidades  #################

    # Servicios afectados por vulnerabilidades criticas
    #servicios_vuln_criticas = c.execute("select COUNT (IP) from VULNERABILIDADES where tipo ='ms17010' or tipo ='ms08067' or tipo ='passTomcat'  or tipo ='zimbraXXE' or tipo ='doublepulsar' or tipo ='webshell' or tipo ='backdoorFabrica'  or tipo ='ransomware' or tipo ='passwordSFI' or tipo ='BlueKeep' or tipo ='BlueKeep'  or tipo ='JoomlaJCKeditor' or tipo ='zerologon'").fetchone()[0]
    #total_servicios_vuln_criticas = total_servicios_vuln_criticas + servicios_vuln_criticas
    #print(f"total_servicios_vuln_criticas {total_servicios_vuln_criticas}")

    # Servicios afectados por vulnerabilidades altas
    #servicios_vuln_altas = c.execute("select COUNT (IP) from VULNERABILIDADES where tipo ='archivosPeligrosos' or tipo ='mailPass' or tipo ='passwordDefecto' or tipo ='compartidoNFS' or tipo ='compartidoSMB' or tipo ='passwordHost' or tipo ='logeoRemoto' or tipo ='heartbleed' or tipo ='adminPassword' or tipo ='rootPassword' or tipo ='ciscoPassword' or tipo ='passwordMikroTik' or tipo ='VNCnopass' or tipo ='VNCbypass' or tipo ='vulnDahua' or tipo ='openrelay' or tipo ='perdidaAutenticacion' or tipo ='spoof' or tipo ='slowloris' or tipo ='wordpressPass' or tipo ='conficker' or tipo ='anonymousIPMI' or tipo ='noSQLDatabases' or tipo ='winboxVuln' or tipo ='rmiVuln' or tipo ='SSHBypass' or tipo ='intelVuln' or tipo ='backupWeb' or tipo ='apacheStruts' or tipo ='webdavVulnerable' or tipo ='IISwebdavVulnerable' or tipo ='shellshock' or tipo ='ciscoASAVuln' or tipo ='SambaCry' or tipo ='misfortune' or tipo ='jbossVuln' or tipo ='passwordBD' or tipo ='passwordAdivinado' or tipo ='wordpressDesactualizado' or tipo ='llmnr' or tipo ='poisoning' or tipo ='cipherZeroIPMI' or tipo ='vlanHop'  or tipo ='owaVul'  or tipo ='phishing' or tipo ='hashRoto' or tipo ='pluginDesactualizado' or tipo ='cmsDesactualizado'").fetchone()[0]
    #total_servicios_vuln_altas = total_servicios_vuln_altas + servicios_vuln_altas
    #print(f"total_servicios_vuln_altas {total_servicios_vuln_altas}")

    # Servicios afectados por vulnerabilidades medias
    #servicios_vuln_medias = c.execute("select COUNT (IP) from VULNERABILIDADES where tipo ='VPNhandshake' or tipo ='openstreaming' or tipo ='directorioLDAP' or tipo ='enum4linux' or tipo ='transferenciaDNS' or tipo ='listadoDirectorio'  or tipo ='enumeracionUsuarios' or tipo ='googlehacking' or tipo ='anonymous' or tipo ='ACL' or tipo ='openresolver' or tipo ='ms12020' or tipo ='HTTPsys' or tipo ='upnpAbierto' or tipo ='registroHabilitado' or tipo ='stp' or tipo ='noTLS' or tipo ='wordpressPingbacks' or tipo ='vulTLS'").fetchone()[0]
    #total_servicios_vuln_medias  = total_servicios_vuln_medias  + servicios_vuln_medias
    #print(f"total_servicios_vuln_medias  {total_servicios_vuln_medias }")

    # Servicios afectados por vulnerabilidades bajas
    #servicios_vuln_bajas = c.execute("select COUNT (IP) from VULNERABILIDADES where tipo ='passwordDahua' or tipo ='divulgacionInformacion'  or tipo ='snmpCommunity' or tipo ='listadoDirectorio' or tipo ='enumeracionUsuarios' or tipo ='googlehacking' or tipo ='anonymous' or tipo ='erroresWeb' or tipo ='ACL' or tipo ='archivosDefecto' or tipo ='openresolver' or tipo ='listadoDirectorios' or tipo ='CVE15473' or tipo ='IPinterna' or tipo ='ftpAnonymous' or tipo ='contenidoNoRelacionado'  or tipo ='confTLS' or tipo ='captcha' or tipo ='contenidoIndebido' or tipo ='debugHabilitado' ").fetchone()[0]
    #total_servicios_vuln_bajas  = total_servicios_vuln_bajas  + servicios_vuln_bajas
    #print(f"total_servicios_vuln_bajas  {total_servicios_vuln_bajas }")
    ###############################################


    ###### VULNERABILIDADES POR ACTIVO ####

    # aplicacionWeb
    vuln_app = c.execute("SELECT count (distinct tipo) FROM VULNERABILIDADES WHERE TIPO='debugHabilitado' or TIPO='listadoDirectorios' or TIPO='archivosDefecto' or TIPO='divulgacionInformacion' or TIPO='archivosPeligrosos' or TIPO='googlehacking' or TIPO='perdidaAutenticacion' or TIPO='erroresWeb' or TIPO='wpusers' or TIPO='exposicionUsuarios'  or TIPO='wordpressPass' or TIPO='IPinterna' or TIPO='webshell'  or TIPO='backupWeb' or TIPO='wordpressDesactualizado' or TIPO='registroHabilitado' or TIPO='captcha' or tipo ='contenidoIndebido' or tipo ='wordpressPingbacks' or tipo ='pluginDesactualizado' or tipo ='cmsDesactualizado'").fetchone()[0]
    aplicacionWeb   = aplicacionWeb   + vuln_app
    print(f"aplicacionWeb {aplicacionWeb}")

    # servidores
    vuln_serv = c.execute("SELECT count (distinct tipo) FROM VULNERABILIDADES WHERE TIPO='passwordHost' or TIPO='compartidoNFS' or TIPO='enum4linux' or TIPO='shellshock' or TIPO='webdavVulnerable' or TIPO='heartbleed' or TIPO='zimbraXXE' or TIPO='slowloris' or TIPO='CVE15473' or TIPO='directorioLDAP' or TIPO='transferenciaDNS' or TIPO='vrfyHabilitado' or TIPO='openresolver' or TIPO='openrelay' or TIPO='spoof' or TIPO='openrelay2' or TIPO='anonymousIPMI' or TIPO='rmiVuln' or TIPO='SSHBypass' or TIPO='intelVuln' or TIPO='HTTPsys' or TIPO='apacheStruts' or TIPO='IISwebdavVulnerable' or TIPO='SambaCry' or TIPO='jbossVuln' or TIPO='passwordSFI' or TIPO='contenidoNoRelacionado' or TIPO='cipherZeroIPMI' or TIPO='passwordAdivinado' or TIPO='zerologon' or TIPO='vulTLS' or TIPO='owaVul' or TIPO='confTLS' or tipo ='noTLS'  or tipo ='ftpAnonymous' ").fetchone()[0]
    servidores   = servidores   + vuln_serv
    print(f"servidores  {servidores}")

    # base de datos
    vuln_bd = c.execute("SELECT count (distinct tipo) FROM VULNERABILIDADES WHERE TIPO='passwordBD' or TIPO='noSQLDatabases' or TIPO='JoomlaJCKeditor'").fetchone()[0]
    baseDatos   = baseDatos   + vuln_bd
    print(f"baseDatos  {baseDatos}")

    # estaciones de trabajo
    vuln_estacion = c.execute("SELECT count (distinct tipo) FROM VULNERABILIDADES WHERE  TIPO='compartidoSMB' or TIPO='ms17010' or TIPO='ms08067' or TIPO='BlueKeep' or TIPO='ms12020' or TIPO='doublepulsar' or TIPO='conficker' or TIPO='VNCbypass' or TIPO='VNCnopass' or TIPO='ransomware'  or tipo ='llmnr'").fetchone()[0]
    estacionesTrabajo    = estacionesTrabajo   + vuln_estacion
    print(f"estacionesTrabajo   {estacionesTrabajo }")

    # sistema Vigilancia
    vuln_vigilancia = c.execute("SELECT count (distinct tipo) FROM VULNERABILIDADES WHERE TIPO='vulnDahua' or TIPO='openstreaming' or TIPO='passwordDahua'").fetchone()[0]
    sistemaVigilancia    = sistemaVigilancia   + vuln_vigilancia
    print(f"sistemaVigilancia   {sistemaVigilancia }")

    # Dispositivos de red
    vuln_red = c.execute("SELECT count (distinct tipo) FROM VULNERABILIDADES WHERE TIPO='passwordMikroTik' or TIPO='winboxVuln' or TIPO='passwordDefecto' or TIPO='snmpCommunity' or TIPO='VPNhandshake' or TIPO='backdoorFabrica' or TIPO='ciscoASAVuln' or TIPO='misfortune' or TIPO='upnpAbierto' or TIPO='poisoning' or TIPO='stp' or TIPO='vlanHop'").fetchone()[0]
    dispositivosRed    = dispositivosRed   + vuln_red
    print(f"dispositivosRed   {dispositivosRed }")

    # personal
    vuln_personal = c.execute("SELECT count (distinct tipo) FROM VULNERABILIDADES WHERE TIPO='mailPass' or tipo ='phishing' or tipo ='hashRoto'").fetchone()[0]
    personal = personal   + vuln_personal
    print(f"personal  {personal }")



    #### VULNERABILIDADES POR CATEGORIA ####
    # password
    vuln_pass = c.execute("SELECT count (distinct tipo) FROM VULNERABILIDADES WHERE TIPO='passwordMikroTik' or TIPO='passwordAdivinado' or TIPO='passwordHost' or TIPO='passwordDefecto' or TIPO='mailPass' or TIPO='passwordDahua' or TIPO='passwordBD' or TIPO='noSQLDatabases' or TIPO='passwordSFI' or tipo ='hashRoto'").fetchone()[0]
    passwordDebil = passwordDebil + vuln_pass
    print(f"passwordDebil  {passwordDebil}")

    # falta de parches
    vuln_parches = c.execute("SELECT count (distinct tipo) FROM VULNERABILIDADES WHERE TIPO='winboxVuln' or TIPO='shellshock' or TIPO='ms17010' or TIPO='ms08067' or TIPO='heartbleed' or TIPO='zimbraXXE' or TIPO='BlueKeep' or TIPO='slowloris' or TIPO='CVE15473' or TIPO='ms12020' or TIPO='vulnDahua' or TIPO='webdavVulnerable' or TIPO='doublepulsar' or TIPO='conficker' or TIPO='SSHBypass' or TIPO='VNCbypass' or TIPO='intelVuln' or TIPO='HTTPsys' or TIPO='apacheStruts' or TIPO='backdoorFabrica' or TIPO='IISwebdavVulnerable' or TIPO='ciscoASAVuln' or TIPO='SambaCry' or TIPO='misfortune' or TIPO='jbossVuln' or TIPO='cipherZeroIPMI' or TIPO='ransomware' or TIPO='JoomlaJCKeditor' or TIPO='zerologon' or TIPO='owaVul' or TIPO='pluginDesactualizado' or TIPO='cmsDesactualizado' ").fetchone()[0]
    faltaParches = faltaParches + vuln_parches
    print(f"faltaParches  {faltaParches}")

    # Errores de configuracion
    vuln_conf = c.execute("SELECT count (distinct tipo) FROM VULNERABILIDADES WHERE TIPO='logeoRemoto' or TIPO='compartidoNFS' or TIPO='compartidoSMB' or TIPO='enum4linux' or TIPO='snmpCommunity' or TIPO='directorioLDAP' or TIPO='transferenciaDNS' or TIPO='vrfyHabilitado' or TIPO='ftpAnonymous' or TIPO='openstreaming' or TIPO='VPNhandshake' or TIPO='openresolver' or TIPO='openrelay' or TIPO='spoof' or TIPO='anonymousIPMI' or TIPO='rmiVuln' or TIPO='VNCnopass' or TIPO='upnpAbierto' or TIPO='contenidoNoRelacionado' or TIPO ='llmnr' or TIPO ='poisoning' or TIPO ='registroHabilitado' or TIPO='stp' or TIPO='vlanHop' or TIPO='vulTLS' or TIPO='confTLS'  or TIPO='captcha' or tipo ='noTLS' or tipo ='contenidoIndebido' or tipo ='wordpressPingbacks' or tipo ='archivosDefecto' or tipo ='listadoDirectorios' or tipo ='archivosPeligrosos' or tipo ='erroresWeb' or tipo ='debugHabilitado'").fetchone()[0]
    errorConfiguracion = errorConfiguracion + vuln_conf
    print(f"errorConfiguracion  {errorConfiguracion}")





    ################## CREAR INFORME ############################
    for empresa in empresas:
        print (f"Generando informe para {empresa}")
        # crear hoja de calculo
        globals()['wb' + empresa] = openpyxl.Workbook()


        ###### ESTADISTICAS INTERNAS/EXTERNAS ####
        # crear nueva pestania
        globals()['wb' + empresa].create_sheet()
        globals()['sheet' + empresa] = globals()['wb' + empresa]['Sheet']

        globals()['sheet' + empresa].title = 'estadisticas'  # Change title

        globals()['sheet' + empresa].column_dimensions['A'].width = 20
        globals()['sheet' + empresa].column_dimensions['B'].width = 25
        globals()['sheet' + empresa]['A1'] = "Valor"
        globals()['sheet' + empresa]['A2'] = "Crítico"
        globals()['sheet' + empresa]['A3'] = "Alto"
        globals()['sheet' + empresa]['A4'] = "Medio"
        globals()['sheet' + empresa]['A5'] = "Bajo"
        #globals()['sheet' + empresa]['A6'] = "Informativo"

        globals()['sheet' + empresa]['A1'].border = thin_border
        globals()['sheet' + empresa]['A2'].border = thin_border
        globals()['sheet' + empresa]['A3'].border = thin_border
        globals()['sheet' + empresa]['A4'].border = thin_border
        globals()['sheet' + empresa]['A5'].border = thin_border
        #globals()['sheet' + empresa]['A6'].border = thin_border

        globals()['sheet' + empresa]['B1'].border = thin_border
        globals()['sheet' + empresa]['B2'].border = thin_border
        globals()['sheet' + empresa]['B3'].border = thin_border
        globals()['sheet' + empresa]['B4'].border = thin_border
        globals()['sheet' + empresa]['B5'].border = thin_border
        #globals()['sheet' + empresa]['B6'].border = thin_border

        globals()['sheet' + empresa]['A1'].alignment = Alignment(horizontal="center", vertical='center', wrap_text=True)
        globals()['sheet' + empresa]['A2'].alignment = Alignment(horizontal="center", vertical='center', wrap_text=True)
        globals()['sheet' + empresa]['A3'].alignment = Alignment(horizontal="center", vertical='center', wrap_text=True)
        globals()['sheet' + empresa]['A4'].alignment = Alignment(horizontal="center", vertical='center', wrap_text=True)
        globals()['sheet' + empresa]['A5'].alignment = Alignment(horizontal="center", vertical='center', wrap_text=True)
        #globals()['sheet' + empresa]['A6'].alignment = Alignment(horizontal="center", vertical='center', wrap_text=True)

        globals()['sheet' + empresa]['B1'].alignment = Alignment(horizontal="center", vertical='center', wrap_text=True)
        globals()['sheet' + empresa]['B2'].alignment = Alignment(horizontal="center", vertical='center', wrap_text=True)
        globals()['sheet' + empresa]['B3'].alignment = Alignment(horizontal="center", vertical='center', wrap_text=True)
        globals()['sheet' + empresa]['B4'].alignment = Alignment(horizontal="center", vertical='center', wrap_text=True)
        globals()['sheet' + empresa]['B5'].alignment = Alignment(horizontal="center", vertical='center', wrap_text=True)
        #globals()['sheet' + empresa]['B6'].alignment = Alignment(horizontal="center", vertical='center', wrap_text=True)

        globals()['sheet' + empresa]['A1'].fill = negroFill
        globals()['sheet' + empresa]['A1'].font = Arial12BoldWhite
        globals()['sheet' + empresa]['A2'].fill = criticoFill
        globals()['sheet' + empresa]['A3'].fill = AltoFill2
        globals()['sheet' + empresa]['A4'].fill = ModeradoFill
        globals()['sheet' + empresa]['A5'].fill = bajoFill2
        #globals()['sheet' + empresa]['A6'].fill = MuyBajoFill

        globals()['sheet' + empresa]['B1'].fill = negroFill
        globals()['sheet' + empresa]['B1'].font = Arial12BoldWhite
        globals()['sheet' + empresa]['B1'] = "Número de riesgos"
        globals()['sheet' + empresa]['B2'] = globals()['total_vul_criticas_' + vectorInforme]
        globals()['sheet' + empresa]['B3'] = globals()['total_vul_altas_' + vectorInforme]
        globals()['sheet' + empresa]['B4'] = globals()['total_vul_medias_' + vectorInforme]
        globals()['sheet' + empresa]['B5'] = globals()['total_vul_bajas_' + vectorInforme]
        #globals()['sheet' + empresa]['B6'] = 0

        chart = PieChart3D()

        # create data for plotting
        labels = Reference( globals()['sheet' + empresa], min_col=1, min_row=2, max_row=5)
        data = Reference( globals()['sheet' + empresa], min_col=2, min_row=2, max_row=5)

        # adding data to the Doughnut chart object
        chart.add_data(data, titles_from_data=False)
        chart.set_categories(labels)

        chart.dataLabels = DataLabelList()
        chart.dataLabels.showPercent = True
        chart.dataLabels.showVal = False
        chart.dataLabels.showLegendKey = False
        chart.dataLabels.showCatName = False

        # set style of the chart
        chart.style = 26

        # try to set color blue (0000FF) for the 2nd wedge (idx=1) in the series
        series = chart.series[0]
        pt = DataPoint(idx=0)
        pt.graphicalProperties.solidFill = "FF0000"
        series.dPt.append(pt)

        pt = DataPoint(idx=1)
        pt.graphicalProperties.solidFill = "FFC000"
        series.dPt.append(pt)

        pt = DataPoint(idx=2)
        pt.graphicalProperties.solidFill = "FFFF00"
        series.dPt.append(pt)

        pt = DataPoint(idx=3)
        pt.graphicalProperties.solidFill = "92D050"
        series.dPt.append(pt)

        pt = DataPoint(idx=4)
        pt.graphicalProperties.solidFill = "95B3D7"
        series.dPt.append(pt)
        #adicionar la grafica a la hoja de calculo
        globals()['sheet' + empresa].add_chart(chart, 'C5')




        ###### PUERTOS ABIERTOS ####
        nmapDF = pd.read_csv(ruta + "reportes/NMAP-resumen.txt", sep='\t', encoding='utf-8', header=None)
        # print(nmapDF.head())
        globals()['wb' + empresa].create_sheet()
        globals()['sheet' + empresa] = globals()['wb' + empresa]['Sheet']
        globals()['sheet' + empresa].title = 'Puertos abiertos'  # Change title
        globals()['sheet' + empresa].column_dimensions['A'].width = 15
        globals()['sheet' + empresa].column_dimensions['B'].width = 25
        globals()['sheet' + empresa].column_dimensions['C'].width = 15

        i = 1
        # Cabecera
        i = i + 1
        globals()['sheet' + empresa].merge_cells('B' + str(i) + ':C' + str(i))
        globals()['sheet' + empresa]['A' + str(i)] = "IP"
        globals()['sheet' + empresa]['A' + str(i)].border = thin_border
        globals()['sheet' + empresa]['A' + str(i)].alignment = Alignment(horizontal="center", vertical='center', wrap_text=True)
        globals()['sheet' + empresa]['A' + str(i)].font = Calibri10BoldWhite
        globals()['sheet' + empresa]['A' + str(i)].fill = verdeFill

        globals()['sheet' + empresa]['B' + str(i)] = "Puertos abiertos"
        globals()['sheet' + empresa]['B' + str(i)].border = thin_border
        globals()['sheet' + empresa]['C' + str(i)].border = thin_border
        globals()['sheet' + empresa]['B' + str(i)].alignment = Alignment(horizontal="center", vertical='center', wrap_text=True)
        globals()['sheet' + empresa]['B' + str(i)].font = Calibri10BoldWhite
        globals()['sheet' + empresa]['B' + str(i)].fill = verdeFill

        i = i + 1
        for index, row in nmapDF.iterrows():
            ip = row[0]
            tcp = row[1]
            udp = row[2]

            globals()['sheet' + empresa]['A' + str(i)] = ip
            globals()['sheet' + empresa]['B' + str(i)] = tcp
            globals()['sheet' + empresa]['C' + str(i)] = udp

            globals()['sheet' + empresa]['A' + str(i)].font = Calibri10
            globals()['sheet' + empresa]['B' + str(i)].font = Calibri10
            globals()['sheet' + empresa]['C' + str(i)].font = Calibri10

            globals()['sheet' + empresa]['A' + str(i)].alignment = Alignment(horizontal="center", vertical='center', wrap_text=True)
            globals()['sheet' + empresa]['B' + str(i)].alignment = Alignment(horizontal="center", vertical='center', wrap_text=True)
            globals()['sheet' + empresa]['C' + str(i)].alignment = Alignment(horizontal="center", vertical='center', wrap_text=True)

            globals()['sheet' + empresa]['A' + str(i)].border = thin_border
            globals()['sheet' + empresa]['B' + str(i)].border = thin_border
            globals()['sheet' + empresa]['C' + str(i)].border = thin_border
            i = i + 1

        ###### SISTEMAS OPERATIVOS ####
        if (os.path.isfile(ruta + "reportes/reporte-OS.csv")):
            osDF = pd.read_csv(ruta + "reportes/reporte-OS.csv", sep=';', encoding='utf-8')
            # print(nmapDF.head())
            globals()['wb' + empresa].create_sheet()
            globals()['sheet' + empresa] = globals()['wb' + empresa]['Sheet']
            globals()['sheet' + empresa].title = 'Sistemas operativos'  # Change title
            globals()['sheet' + empresa].column_dimensions['A'].width = 15
            globals()['sheet' + empresa].column_dimensions['B'].width = 15
            globals()['sheet' + empresa].column_dimensions['C'].width = 15
            globals()['sheet' + empresa].column_dimensions['D'].width = 30

            i = 1
            # Cabecera
            i = i + 1
            globals()['sheet' + empresa]['A' + str(i)] = "IP"
            globals()['sheet' + empresa]['A' + str(i)].border = thin_border
            globals()['sheet' + empresa]['A' + str(i)].alignment = Alignment(horizontal="center", vertical='center', wrap_text=True)
            globals()['sheet' + empresa]['A' + str(i)].font = Calibri10BoldWhite
            globals()['sheet' + empresa]['A' + str(i)].fill = verdeFill

            globals()['sheet' + empresa]['B' + str(i)] = "Nombre"
            globals()['sheet' + empresa]['B' + str(i)].border = thin_border
            globals()['sheet' + empresa]['B' + str(i)].alignment = Alignment(horizontal="center", vertical='center', wrap_text=True)
            globals()['sheet' + empresa]['B' + str(i)].font = Calibri10BoldWhite
            globals()['sheet' + empresa]['B' + str(i)].fill = verdeFill

            globals()['sheet' + empresa]['C' + str(i)] = "Dominio"
            globals()['sheet' + empresa]['C' + str(i)].border = thin_border
            globals()['sheet' + empresa]['C' + str(i)].alignment = Alignment(horizontal="center", vertical='center', wrap_text=True)
            globals()['sheet' + empresa]['C' + str(i)].font = Calibri10BoldWhite
            globals()['sheet' + empresa]['C' + str(i)].fill = verdeFill

            globals()['sheet' + empresa]['D' + str(i)] = "S.O."
            globals()['sheet' + empresa]['D' + str(i)].border = thin_border
            globals()['sheet' + empresa]['D' + str(i)].alignment = Alignment(horizontal="center", vertical='center', wrap_text=True)
            globals()['sheet' + empresa]['D' + str(i)].font = Calibri10BoldWhite
            globals()['sheet' + empresa]['D' + str(i)].fill = verdeFill

            i = i + 1
            for index, row in osDF.iterrows():
                ip = row[0]
                nombre = row[1]
                dominio = row[2]
                sistemaOperativo = str(row[4])
                # print(f"os {sistemaOperativo}")
                # print(f"os type {type(sistemaOperativo)}")
                if sistemaOperativo == "nan":
                    continue

                globals()['sheet' + empresa]['A' + str(i)] = ip
                globals()['sheet' + empresa]['B' + str(i)] = nombre
                globals()['sheet' + empresa]['C' + str(i)] = dominio
                globals()['sheet' + empresa]['D' + str(i)] = sistemaOperativo

                globals()['sheet' + empresa]['A' + str(i)].font = Calibri10
                globals()['sheet' + empresa]['B' + str(i)].font = Calibri10
                globals()['sheet' + empresa]['C' + str(i)].font = Calibri10
                globals()['sheet' + empresa]['D' + str(i)].font = Calibri10

                globals()['sheet' + empresa]['A' + str(i)].alignment = Alignment(horizontal="center", vertical='center', wrap_text=True)
                globals()['sheet' + empresa]['B' + str(i)].alignment = Alignment(horizontal="center", vertical='center', wrap_text=True)
                globals()['sheet' + empresa]['C' + str(i)].alignment = Alignment(horizontal="center", vertical='center', wrap_text=True)
                globals()['sheet' + empresa]['D' + str(i)].alignment = Alignment(horizontal="center", vertical='center', wrap_text=True)

                globals()['sheet' + empresa]['A' + str(i)].border = thin_border
                globals()['sheet' + empresa]['B' + str(i)].border = thin_border
                globals()['sheet' + empresa]['C' + str(i)].border = thin_border
                globals()['sheet' + empresa]['D' + str(i)].border = thin_border
                i = i + 1

        ###### SUBDOMINIOS ####
        if (os.path.isfile(ruta + "importarMaltego/subdominios.csv")):
            subdomainDF = pd.read_csv(ruta + "importarMaltego/subdominios.csv", sep=',', encoding='utf-8', header=None)
            # print(nmapDF.head())
            globals()['wb' + empresa].create_sheet()
            globals()['sheet' + empresa] = globals()['wb' + empresa]['Sheet']
            globals()['sheet' + empresa].title = 'Subdominios'  # Change title
            globals()['sheet' + empresa].column_dimensions['A'].width = 18
            globals()['sheet' + empresa].column_dimensions['B'].width = 18
            globals()['sheet' + empresa].column_dimensions['C'].width = 18
            globals()['sheet' + empresa].column_dimensions['D'].width = 18

            i = 1
            # Cabecera
            i = i + 1
            globals()['sheet' + empresa]['A' + str(i)] = "IP"
            globals()['sheet' + empresa]['A' + str(i)].border = thin_border
            globals()['sheet' + empresa]['A' + str(i)].alignment = Alignment(horizontal="center", vertical='center', wrap_text=True)
            globals()['sheet' + empresa]['A' + str(i)].font = Calibri10BoldWhite
            globals()['sheet' + empresa]['A' + str(i)].fill = verdeFill

            globals()['sheet' + empresa]['B' + str(i)] = "Subdominio"
            globals()['sheet' + empresa]['B' + str(i)].border = thin_border
            globals()['sheet' + empresa]['B' + str(i)].alignment = Alignment(horizontal="center", vertical='center', wrap_text=True)
            globals()['sheet' + empresa]['B' + str(i)].font = Calibri10BoldWhite
            globals()['sheet' + empresa]['B' + str(i)].fill = verdeFill

            globals()['sheet' + empresa]['C' + str(i)] = "Ubicación"
            globals()['sheet' + empresa]['C' + str(i)].border = thin_border
            globals()['sheet' + empresa]['C' + str(i)].alignment = Alignment(horizontal="center", vertical='center', wrap_text=True)
            globals()['sheet' + empresa]['C' + str(i)].font = Calibri10BoldWhite
            globals()['sheet' + empresa]['C' + str(i)].fill = verdeFill

            globals()['sheet' + empresa]['D' + str(i)] = "ISP"
            globals()['sheet' + empresa]['D' + str(i)].border = thin_border
            globals()['sheet' + empresa]['D' + str(i)].alignment = Alignment(horizontal="center", vertical='center', wrap_text=True)
            globals()['sheet' + empresa]['D' + str(i)].font = Calibri10BoldWhite
            globals()['sheet' + empresa]['D' + str(i)].fill = verdeFill

            i = i + 1
            for index, row in subdomainDF.iterrows():
                ip = row[1]
                subdominio = row[2]
                ubicacion = row[3]
                isp = str(row[4])
                # print(f"os {sistemaOperativo}")
                # print(f"os type {type(sistemaOperativo)}")

                globals()['sheet' + empresa]['A' + str(i)] = ip
                globals()['sheet' + empresa]['B' + str(i)] = subdominio
                globals()['sheet' + empresa]['C' + str(i)] = ubicacion
                globals()['sheet' + empresa]['D' + str(i)] = isp

                globals()['sheet' + empresa]['A' + str(i)].font = Calibri10
                globals()['sheet' + empresa]['B' + str(i)].font = Calibri10
                globals()['sheet' + empresa]['C' + str(i)].font = Calibri10
                globals()['sheet' + empresa]['D' + str(i)].font = Calibri10

                globals()['sheet' + empresa]['A' + str(i)].alignment = Alignment(horizontal="center", vertical='center', wrap_text=True)
                globals()['sheet' + empresa]['B' + str(i)].alignment = Alignment(horizontal="center", vertical='center', wrap_text=True)
                globals()['sheet' + empresa]['C' + str(i)].alignment = Alignment(horizontal="center", vertical='center', wrap_text=True)
                globals()['sheet' + empresa]['D' + str(i)].alignment = Alignment(horizontal="center", vertical='center', wrap_text=True)

                globals()['sheet' + empresa]['A' + str(i)].border = thin_border
                globals()['sheet' + empresa]['B' + str(i)].border = thin_border
                globals()['sheet' + empresa]['C' + str(i)].border = thin_border
                globals()['sheet' + empresa]['D' + str(i)].border = thin_border
                i = i + 1

        ##### PRUEBAS REALIZADAS ####
        print("Generando reporte de pruebas realizadas")
        globals()['wb' + empresa].create_sheet()
        globals()['sheet' + empresa] = globals()['wb' + empresa]['Sheet']
        globals()['sheet' + empresa].title = 'pruebas'  # Change title.
        globals()['sheet' + empresa].column_dimensions['A'].width = 15
        globals()['sheet' + empresa].column_dimensions['B'].width = 10
        globals()['sheet' + empresa].column_dimensions['C'].width = 15
        globals()['sheet' + empresa].column_dimensions['D'].width = 48

        # Num pruebas de vulnerabilidades
        stream = os.popen('ls ' + ruta + '/logs/vulnerabilidades | cut -d "_" -f3 | sort | uniq | wc -l')
        pruebasVulnerabilidades_cant = int(stream.read())

        # Num pruebas de cracking
        stream = os.popen('ls ' + ruta + '/logs/cracking | cut -d "_" -f3 | sort | uniq | wc -l')
        pruebasPassword_cant = int(stream.read())

        totalPruebas = totalPruebas + pruebasVulnerabilidades_cant + pruebasPassword_cant
        print(f"totalPruebas {totalPruebas}")

        # pruebas de vulnerabilidades
        stream = os.popen('ls ' + ruta + '/logs/vulnerabilidades')
        pruebasVulnerabilidades = stream.read()
        pruebasVulnerabilidades_list = pruebasVulnerabilidades.split('\n')
        pruebasVulnerabilidades_list.pop()
        # print(f"pruebasVulnerabilidades {pruebasVulnerabilidades_list}")

        # pruebas de Cracking
        stream = os.popen('ls ' + ruta + '/logs/cracking')
        pruebasCracking = stream.read()
        pruebasCracking_list = pruebasCracking.split('\n')
        pruebasCracking_list.pop()
        # print(f"pruebasCracking {pruebasCracking_list}")

        # pruebas de Enumeracion
        stream = os.popen('ls ' + ruta + '/logs/enumeracion/ | grep --color=never web | egrep -v "Data|canary|wget"')
        pruebasEnumeracion = stream.read()
        pruebasEnumeracion_list = pruebasEnumeracion.split('\n')
        pruebasEnumeracion_list.pop()
        # print(f"pruebasEnumeracion {pruebasEnumeracion_list}")

        todasPruebas = pruebasVulnerabilidades_list + pruebasCracking_list + pruebasEnumeracion_list
        #print(f"todasPruebas {todasPruebas}")

        stream = os.popen('cat ' + ruta + 'top.txt | wc -l')
        totalPasswords = int(stream.read())  # totalPasswords probados
        # print(f"totalPasswords {totalPasswords}")

        webApp_tests = []
        servers_tests = []
        bd_tests = []
        workstation_tests = []
        video_tests = []
        network_tests = []
        passwords_tests = []
        testHacking = []

        for prueba in todasPruebas:
#            print(f"prueba {prueba}")
            prueba_list = prueba.split('_')
            ip = prueba_list[0]
            port = prueba_list[1]
            vuln = prueba_list[2]
            vuln = vuln.replace(".txt", "")
            vuln = vuln.replace(".html", "")
            cod_desc = buscarDescripcion(vuln)
            codVul = cod_desc.split(";")[0]
            desc = cod_desc.split(";")[1]

            desc = desc.replace("DOMINIO", args.domain)
            # print(f"vuln {vuln} ip {ip} port {port} codVul {codVul} desc {desc}")

            # Activos de informacion - aplicaciones web
            if vuln in ['debugHabilitado', 'listadoDirectorios', 'divulgacionInformacion', 'archivosDefecto',
                        'archivosPeligrosos', 'erroresWeb', 'wpusers', 'perdidaAutenticacion', 'exposicionUsuarios',
                        'wordpressPass', 'IPinterna', 'backupWeb', 'pluginDesactualizado', 'webshell',
                        'registroHabilitado',
                        'contenidoIndebido', 'JoomlaJCKeditor','cmsDesactualizado']:
                webApp_tests.append({'ip': ip, 'port': port, 'codVul': codVul, 'desc': desc})
            if 'googlehacking' in vuln:
                webApp_tests.append({'ip': ip, 'port': "N/A", 'codVul': codVul, 'desc': desc})

            if 'password' in vuln:
                passwords_tests.append({'ip': ip, 'port': port, 'codVul': codVul, 'desc': desc})
            # aplicacionWeb_csv = aplicacionWeb_csv."$ip~$port~$codVul~vuln_descripcion~vuln\n";

            # Activos de informacion - servidores
            if vuln in ['compartidoNFS', 'enum4linux', 'shellshock', 'webdavVulnerable', 'heartbleed', 'zimbraXXE',
                        'slowloris', 'CVE15473', 'directorioLDAP', 'transferenciaDNS', 'vrfyHabilitado', 'openresolver',
                        'openrelay', 'anonymousIPMI', 'rmiVuln', 'SSHBypass', 'intelVuln', 'HTTPsys', 'apacheStruts',
                        'IISwebdavVulnerable', 'SambaCry', 'jbossVuln', 'contenidoNoRelacionado', 'spoof',
                        'cipherZeroIPMI',
                        'zerologon', 'vulTLS', 'confTLS', 'owaVul', 'noTLS', 'ftpAnonymous','passwordHost']:
                servers_tests.append({'ip': ip, 'port': port, 'codVul': codVul, 'desc': desc})

            # Activos de informacion - baseDatos
            if vuln in ['noSQLDatabases']:
                bd_tests.append({'ip': ip, 'port': port, 'codVul': codVul, 'desc': desc})

            # Activos de informacion - estacionesTrabajo
            if vuln in ['compartidoSMB', 'ms17010', 'ms08067', 'BlueKeep', 'ms12020', 'doublepulsar', 'conficker',
                        'VNCbypass', 'VNCnopass', 'ransomware', 'llmnr', 'hashRoto']:
                workstation_tests.append({'ip': ip, 'port': port, 'codVul': codVul, 'desc': desc})

            # Activos de informacion - sistemaVigilancia
            if vuln in ['vulnDahua', 'openstreaming', 'passwordDahua']:
                video_tests.append({'ip': ip, 'port': port, 'codVul': codVul, 'desc': desc})

            # Activos de informacion - dispositivosRed
            if vuln in ['winboxVuln', 'snmpCommunity', 'VPNhandshake', 'backdoorFabrica', 'ciscoASAVuln', 'misfortune',
                        'upnpAbierto', 'poisoning', 'stp', 'vlanHop']:
                network_tests.append({'ip': ip, 'port': port, 'codVul': codVul, 'desc': desc})

        testHacking.append({'title': "Pruebas a aplicaciones web", 'tests': webApp_tests})
        testHacking.append({'title': "Pruebas a servidores (web, SMB, correo, etc)", 'tests': servers_tests})
        testHacking.append({'title': "Pruebas a base de datos", 'tests': bd_tests})
        testHacking.append({'title': "Pruebas a estaciones de trabajo", 'tests': workstation_tests})
        testHacking.append({'title': "Pruebas a sistemas de vigilancia ", 'tests': video_tests})
        testHacking.append({'title': "Pruebas a dispositivos de red", 'tests': network_tests})
        testHacking.append(
            {'title': f"Pruebas de password a servicios y dispositivos (Passwords probados {totalPasswords})",
             'tests': passwords_tests})

        i = 1
        for details in testHacking:
            testTitle = details['title']
            test_category = details['tests']

            # Cabecera
            i = i + 1
            globals()['sheet' + empresa].merge_cells('A' + str(i) + ':D' + str(i))
            globals()['sheet' + empresa]['A' + str(i)] = testTitle
            globals()['sheet' + empresa]['A' + str(i)].border = thin_border
            globals()['sheet' + empresa]['A' + str(i)].alignment = Alignment(horizontal="center", vertical='center', wrap_text=True)
            globals()['sheet' + empresa]['A' + str(i)].font = Calibri10BoldWhite
            globals()['sheet' + empresa]['A' + str(i)].fill = verdeFill
            i = i + 1

            globals()['sheet' + empresa]['A' + str(i)] = "IP"
            globals()['sheet' + empresa]['B' + str(i)] = "Puerto"
            globals()['sheet' + empresa]['C' + str(i)] = "Código CVE"
            globals()['sheet' + empresa]['D' + str(i)] = "Prueba realizada"

            globals()['sheet' + empresa]['A' + str(i)].border = thin_border
            globals()['sheet' + empresa]['B' + str(i)].border = thin_border
            globals()['sheet' + empresa]['C' + str(i)].border = thin_border
            globals()['sheet' + empresa]['D' + str(i)].border = thin_border

            globals()['sheet' + empresa]['A' + str(i)].alignment = Alignment(horizontal="center", vertical='center', wrap_text=True)
            globals()['sheet' + empresa]['B' + str(i)].alignment = Alignment(horizontal="center", vertical='center', wrap_text=True)
            globals()['sheet' + empresa]['C' + str(i)].alignment = Alignment(horizontal="center", vertical='center', wrap_text=True)
            globals()['sheet' + empresa]['D' + str(i)].alignment = Alignment(horizontal="center", vertical='center', wrap_text=True)

            globals()['sheet' + empresa]['A' + str(i)].font = Calibri10BoldWhite
            globals()['sheet' + empresa]['B' + str(i)].font = Calibri10BoldWhite
            globals()['sheet' + empresa]['C' + str(i)].font = Calibri10BoldWhite
            globals()['sheet' + empresa]['D' + str(i)].font = Calibri10BoldWhite

            globals()['sheet' + empresa]['A' + str(i)].fill = verdeFill
            globals()['sheet' + empresa]['B' + str(i)].fill = verdeFill
            globals()['sheet' + empresa]['C' + str(i)].fill = verdeFill
            globals()['sheet' + empresa]['D' + str(i)].fill = verdeFill

            i = i + 1
            for test in test_category:
                ip = test['ip']
                port = test['port']
                codVul = test['codVul']
                desc = test['desc']

                globals()['sheet' + empresa]['A' + str(i)] = ip
                globals()['sheet' + empresa]['B' + str(i)] = port
                globals()['sheet' + empresa]['C' + str(i)] = codVul
                globals()['sheet' + empresa]['D' + str(i)] = desc

                globals()['sheet' + empresa]['A' + str(i)].font = Calibri10
                globals()['sheet' + empresa]['B' + str(i)].font = Calibri10
                globals()['sheet' + empresa]['C' + str(i)].font = Calibri10
                globals()['sheet' + empresa]['D' + str(i)].font = Calibri10

                globals()['sheet' + empresa]['A' + str(i)].alignment = Alignment(horizontal="center", vertical='center', wrap_text=True)
                globals()['sheet' + empresa]['B' + str(i)].alignment = Alignment(horizontal="center", vertical='center', wrap_text=True)
                globals()['sheet' + empresa]['C' + str(i)].alignment = Alignment(horizontal="center", vertical='center', wrap_text=True)
                globals()['sheet' + empresa]['D' + str(i)].alignment = Alignment(horizontal="center", vertical='center', wrap_text=True)

                globals()['sheet' + empresa]['A' + str(i)].border = thin_border
                globals()['sheet' + empresa]['B' + str(i)].border = thin_border
                globals()['sheet' + empresa]['C' + str(i)].border = thin_border
                globals()['sheet' + empresa]['D' + str(i)].border = thin_border

                i = i + 1

        ######################  CREAR FICHAS DE VULNERABILIDADES ############
        numeroVulnerabilidad = 1
        for i in range(total_vulnerabilidades):
            conclusion = ""
            recomendacionMatriz = ""
            cod = root[i].find("cod").text
            nombre = root[i].find("nombre").text
            #print("nombre " + nombre)
            codVul = root[i].find("codVul").text
            #CVSS = root[i].find("CVSS").text
            descripcion = root[i].find("descripcion").text
            detallesTest = root[i].find("detallesTest").text
            activo = root[i].find("activo").text
            detallesTest = detallesTest.replace("DOMINIOENTIDAD", args.domain)
            vector = root[i].find("vector").text
            score = root[i].find("score").text
            riesgoInforme = root[i].find("riesgoInforme").text
            agente_amenaza = root[i].find("agente_amenaza").text
            impacto_tecnico = root[i].find("impacto_tecnico").text
            impacto_negocio = root[i].find("impacto_negocio").text
            referencias = root[i].find("referencias").text
            recomendacion = root[i].find("recomendacion").text

            recomendacion = recomendacion.replace("DOMINIOENTIDAD", args.domain)
            recomendacion = recomendacion.replace("AMPERSAND", '&')
            verificacion = root[i].find("verificacion").text
            try:
                conclusion = root[i].find("conclusion").text
                recomendacionMatriz = root[i].find("recomendacionMatriz").text
                descripcion = descripcion.replace("SALTOLINEA", '\n')
            except:
                error = ""
                #print("No conclu")

            #print (f"cod {cod} conclusion {conclusion}")
            sql = "SELECT * FROM VULNERABILIDADES WHERE TIPO=\"" + cod + "\";"
            resSQL = c.execute(sql)
            filas = 1
            hosts = ""

            # ACCOUNT FOUND: [postgres] Host: 192.168.2.222 User: postgres Password:  [SUCCESS]
            # [MongoDB] $respuesta
            # [Redis] $respuesta"
            if ("passwordBD" in cod):
                for row in resSQL:
                    vuln_detalles = row[3]
                    vuln_detalles = vuln_detalles.replace("User", "Usuario")
                    vuln_detalles = vuln_detalles.replace("ACCOUNT FOUND:", "")
                    vuln_detalles = vuln_detalles.replace("[SUCCESS]", "")
                    vuln_detalles = vuln_detalles.replace("Host", "IP")
                    vuln_detalles = vuln_detalles.replace("Password encontrado:", "")
                    hosts = hosts + vuln_detalles + "\n"
                    filas = filas + 1

            # Solo se muestra la IP y el puerto en el campo "host"
            if cod in ['ms17010', 'ms08067', 'passwordDahua', 'heartbleed', 'directorioLDAP', 'spoof', 'ftpAnonymous',
                       'anonymousIPMI', 'openstreaming', 'VPNhandshake', 'zimbraXXE',
                       'BlueKeep', 'slowloris', 'openresolver', 'openrelay', 'CVE15473', 'ms12020', 'doublepulsar',
                       'conficker',
                       'rmiVuln', 'SSHBypass', 'VNCnopass', 'intelVuln',
                       'HTTPsys', 'apacheStruts', 'IISwebdavVulnerable', 'SambaCry', 'misfortune', 'jbossVuln',
                       'upnpAbierto',
                       'contenidoNoRelacionado', 'ransomware', 'cipherZeroIPMI', 'stp',
                       'vlanHop', 'JoomlaJCKeditor', 'zerologon', 'phishing', 'wordpressPingbacks']:

                for row in resSQL:
                    hosts = hosts + row[0] + "\n"
                    filas = filas + 1
                    # hosts = hosts + row[0] + "\n<br>"
                    # if (( $filas % 5 == 0 ) && ($filas >0))
                # { $hosts = $hosts.$row[0]."\n<br>"; #$hosts = $hosts."<td>".$row[0]."</td></tr><tr>";}
                # else
                # { $hosts = $hosts.$row[0]."&nbsp;&nbsp;&nbsp;"; #$hosts = $hosts."<td>".$row[0]."</td>";}
                # $filas++;

            if ("llmnr" in cod):
                for row in resSQL:
                    ip = row[0]
                    port = row[1]
                    vuln_detalles = row[3].split(':')
                    respuesta_ntlm = ':'.join([vuln_detalles[0], vuln_detalles[1], vuln_detalles[2], vuln_detalles[3],
                                               vuln_detalles[
                                                   4]])  # 'My name is Simon'    `echo '$vuln_detalles' | cut -d ":" -f1-5`;
                    hosts = hosts + ip + ": " + respuesta_ntlm + "\n"
                    filas = filas + 1

            if ("listadoDirectorios" in cod):
                for row in resSQL:
                    ip = row[0]
                    port = row[1]
                    vuln_detalles = row[3]
                    # $vuln_detalles =~ s/\n/<br>- /g;
                    # $vuln_detalles =~ s/-  \n/<br>- /g;
                    if ("index" in vuln_detalles):
                        if port in ['80', '81', '82', '83', '84', '85', '86', '8080', '8081', '8082', '8010',
                                    '8800']:  hosts = hosts + f"- http://{ip}:{port} \n"
                        if port in ['443', '8443', '4443', '4433']:  hosts = hosts + f"- https://{ip}:{port} \n"
                    else:
                        vuln_detalles = vuln_detalles.replace("200 ", "")
                        hosts = hosts + vuln_detalles + "\n"
                    filas = filas + 1

                # $hosts = "- ".$hosts;

            if ("confTLS" in cod):
                for row in resSQL:
                    ip = row[0]
                    port = row[1]
                    vuln_detalles = row[3]
                    vuln_detalles = vuln_detalles.replace("Configuracion incorrecta: ", "")
                    vuln_detalles = vuln_detalles.replace("\n", " ,")

                    hosts = hosts + f"\n-{ip}:{port} : "

                    if ("TLS 1.0 habilitado" in vuln_detalles): hosts = hosts + " TLS 1.0 habilitado, "
                    if ("SSLv3 esta habilitado" in vuln_detalles): hosts = hosts + " SSL 3.0 habilitado, "
                    if ("SSLv2 esta habilitado" in vuln_detalles): hosts = hosts + " SSL 2.0 habilitado, "
                    if ("HSTS" in vuln_detalles): hosts = hosts + " HSTS deshabilitado, "
                    if ("TLS 1.3 no habilitado" in vuln_detalles): hosts = hosts + " TLS 1.3 deshabilitado "
                    filas = filas + 1

            if ("vulTLS" in cod):
                for row in resSQL:
                    ip = row[0]
                    port = row[1]
                    vuln_detalles = row[3]
                    hosts = hosts + f"\n-{ip}:{port}: {vuln_detalles}"
                    filas = filas + 1

            if ("passwordHost" in cod):
                for row in resSQL:
                    ip = row[0]
                    port = row[1]
                    vuln_detalles = row[3]
                    hosts = hosts + f"\n-{ip}:{port} - {vuln_detalles}"
                    filas = filas + 1



            ## Mostrar ip puerto y detalles de la vulnerabilidad
            if cod in ['owaVul', 'cmsDesactualizado']:
                for row in resSQL:
                    ip = row[0]
                    port = row[1]
                    vuln_detalles = "La " + row[3]
                    vuln_detalles = vuln_detalles.replace("VULNERABLE", "es vulnerable")
                    hosts = hosts + f" {ip}:{port} : " + vuln_detalles + "\n"
                    filas = filas + 1

            if cod in ['archivosPeligrosos', 'archivosDefecto', 'perdidaAutenticacion', 'webshell', 'backupWeb',
                       'ciscoASAVuln',
                       'poisoning', 'captcha', 'noTLS', 'contenidoNoRelacionado', 'pluginDesactualizado', 'wpusers','erroresWeb','debugHabilitado']:
                for row in resSQL:
                    ip = row[0]
                    port = row[1]
                    vuln_detalles = row[3]
                    vuln_detalles = vuln_detalles.replace("Mensaje de error", "")
                    vuln_detalles = vuln_detalles.replace("Posible Backdoor", "")
                    vuln_detalles = vuln_detalles.replace("200", "")
                    vuln_detalles = vuln_detalles.replace("TRACE", "")

                    # https://sigec.fonadin.gob.bo:443/.git/	,
                    hosts = hosts + vuln_detalles
                    filas = filas + 1
                # chop($hosts);
                # chop($hosts);

            if ("passwordAdivinado" in cod):
                for row in resSQL:
                    ip = row[0]
                    port = row[1]
                    vuln_detalles = row[3]
                    vuln_detalles = vuln_detalles.replace("Password encontrado:", "")
                    vuln_detalles = vuln_detalles.replace("[FTP] ACCOUNT FOUND:", "")
                    vuln_detalles = vuln_detalles.replace("ACCOUNT FOUND:", "")
                    vuln_detalles = vuln_detalles.replace("[445]", "")
                    vuln_detalles = vuln_detalles.replace("[SUCCESS]", "")
                    vuln_detalles = vuln_detalles.replace("][", "")

                    # [Tomcat] $line (Usuario:tomcat Password:tomcat)
                    # [Cisco] Usuario:cisco $respuesta"
                    # Password encontrado: [PRTG] $url Usuario:$user Password:$password
                    # [AdminWeb] Usuario:admin $respuesta  #401
                    # [445][smb] host: 10.0.0.141   login: administrator   password: Pa$$w0rd
                    # Password encontrado: [Pentaho] $url (Usuario:$user Password:$password)
                    # [FTP] ACCOUNT FOUND: [ftp] Host: 10.0.2.187 User: root Password:  [SUCCESS]
                    # ACCOUNT FOUND: [ftp] Host: 10.0.2.187 User: ftp Password:  [SUCCESS]
                    # [AdminWeb] Usuario:admin 18:38:58 patator    INFO - 200  12563:-1       0.074 | gainza                             |    51 | HTTP/1.1 200 OK

                    hosts = hosts + f" {vuln_detalles}\n"

                    # if($vuln_detalles =~ /Tomcat|Pentaho|AdminWeb/i){$servidores++;}
                    # if($vuln_detalles =~ /Cisco|PRTG/i){$dispositivosRed++;}
                    # if($vuln_detalles =~ /smb/i){$estacionesTrabajo++;}
                    filas = filas + 1

            if ("vrfyHabilitado" in cod):
                for row in resSQL:
                    hosts = hosts + row[0] + "\n"
                    filas = filas + 1

            if ("compartidoSMB" in cod):
                for row in resSQL:
                    ip = row[0]
                    vuln_detalles = row[3]
                    vuln_detalles = vuln_detalles.replace("                                             	", " ")
                    vuln_detalles = vuln_detalles.replace("READ, WRITE", " ")
                    vuln_detalles = vuln_detalles.replace("READ ONLY", " ")
                    hosts = hosts + f"\\\\{ip} {vuln_detalles} "
                    filas = filas + 1

            if ("winboxVuln" in cod):
                for row in resSQL:
                    ip = row[0]
                    vuln_detalles = row[3]
                    vuln_detalles = vuln_detalles.replace(": \n", ":[vacio]")
                    vuln_detalles = vuln_detalles.replace("User", "Usuario")
                    vuln_detalles = vuln_detalles.replace("Pass", "Contraseña")
                    hosts = hosts + f" {ip} (WinBox - MikroTik) - {vuln_detalles} "
                    filas = filas + 1

            if ("snmpCommunity" in cod):
                for row in resSQL:
                    ip = row[0]
                    port = row[1]
                    vuln_detalles = row[3]
                    stream = os.popen('echo "' + vuln_detalles + '" | grep --color=never "Community string"')
                    community_string = stream.read()
                    hosts = hosts + f"{ip}: {community_string} \n"
                    filas = filas + 1

            if ("transferenciaDNS" in cod):
                for row in resSQL:
                    ip = row[0]
                    port = row[1]
                    vuln_detalles = row[3]
                    hosts = hosts + f"{ip}:{port} - Archivo de configuración: {vuln_detalles}\n"
                    filas = filas + 1

            if filas > 1:
                ######## Radical #######
                if (empresa == "Radical"):
                    matrizRecomendaciones.append(
                        {'vectorInforme': vectorInforme, 'activo': activo, 'vulnerabilidad': nombre, 'riesgo': riesgoInforme,
                         'riesgo': riesgoInforme, "recomendacionMatriz": recomendacionMatriz, "conclusion": conclusion})

                    globals()['wb' + empresa].create_sheet()
                    sheet = globals()['wb' + empresa]['Sheet']

                    sheet.column_dimensions['A'].width = 22
                    sheet.column_dimensions['B'].width = 33
                    sheet.column_dimensions['C'].width = 27

                    sheet.title = str(numeroVulnerabilidad)
                    if ("CRÍTICO" in riesgoInforme):
                        sheet['A1'].fill = criticoFill
                        sheet['C5'].fill = criticoFill

                    if ("ALTO" in riesgoInforme):
                        sheet['A1'].fill = altoFill
                        sheet['C5'].fill = altoFill

                    if ("MEDIO" in riesgoInforme):
                        sheet['A1'].fill = medioFill
                        sheet['C5'].fill = medioFill

                    if ("BAJO" in riesgoInforme):
                        sheet['A1'].fill = bajoFill
                        sheet['C5'].fill = bajoFill

                    # Titulo
                    sheet.merge_cells('B1:C1')
                    sheet.row_dimensions[1].height = 59
                    sheet['A1'].font = Impact36
                    sheet['B1'].font = Impact22
                    sheet['A1'].border = thin_border
                    sheet['B1'].border = thin_border
                    sheet['C1'].border = thin_border

                    sheet['A1'].alignment = Alignment(horizontal="center", vertical='center')
                    sheet['B1'].alignment = Alignment(horizontal="center", vertical='center', wrap_text=True)

                    sheet['A1'] = str(numeroVulnerabilidad)
                    sheet['B1'] = str(nombre).upper()

                    # grafico
                    sheet.row_dimensions[2].height = 76.5
                    img = openpyxl.drawing.image.Image('/usr/share/lanscanner/image.png')

                    sheet.add_image(img, "D2")
                    sheet['A2'].border = thin_border
                    sheet['B2'].border = thin_border
                    sheet['C2'].border = thin_border
                    sheet.merge_cells('A2:C2')

                    # texto del grafico
                    sheet.row_dimensions[3].height = 53.2
                    sheet['A3'].font = Arial10
                    sheet['B3'].font = Arial10
                    sheet['C3'].font = Arial10

                    sheet['A3'].alignment = Alignment(horizontal="center", vertical='center', wrap_text=True)
                    sheet['B3'].alignment = Alignment(horizontal="center", vertical='center', wrap_text=True)
                    sheet['C3'].alignment = Alignment(horizontal="center", vertical='center', wrap_text=True)

                    sheet['A3'].border = thin_border
                    sheet['B3'].border = thin_border
                    sheet['C3'].border = thin_border

                    sheet['A3'] = str(agente_amenaza)
                    sheet['B3'] = str(impacto_tecnico)
                    sheet['C3'] = str(impacto_negocio)

                    # ANALISIS DE RIESGO LABEL
                    sheet.merge_cells('A4:C4')
                    sheet.row_dimensions[4].height = 29
                    sheet['A4'].font = Arial11Bold

                    sheet['A4'].border = thin_border
                    sheet['B4'].border = thin_border
                    sheet['C4'].border = thin_border

                    sheet['A4'].alignment = Alignment(horizontal="center", vertical='center', wrap_text=True)
                    sheet['A4'] = "ANALISIS DE RIESGO"
                    sheet['A4'].fill = greyFill

                    # Riesgos
                    sheet.row_dimensions[5].height = 34.5
                    sheet['A5'].font = Arial10
                    sheet['B5'].font = Arial10
                    sheet['C5'].font = Arial11Bold

                    sheet['A5'].alignment = Alignment(horizontal="center", vertical='center', wrap_text=True)
                    sheet['B5'].alignment = Alignment(horizontal="center", vertical='center', wrap_text=True)
                    sheet['C5'].alignment = Alignment(horizontal="center", vertical='center', wrap_text=True)

                    sheet['A5'].border = thin_border
                    sheet['B5'].border = thin_border
                    sheet['C5'].border = thin_border

                    sheet['A5'] = "Score: " + str(score)
                    sheet['B5'] = "Código: " + str(codVul)
                    sheet['C5'] = "RIESGO: " + str(riesgoInforme)

                    # vectpr CVSS
                    sheet.row_dimensions[6].height = 20
                    sheet.merge_cells('B6:C6')
                    sheet['A6'] = "Vector: "
                    sheet['A6'].border = thin_border
                    sheet['A6'].font = Arial11Bold
                    sheet['B6'] =  str(vector)
                    sheet['A6'].alignment = Alignment(horizontal="center", vertical='center', wrap_text=True)
                    sheet['B6'].alignment = Alignment(horizontal="center", vertical='center', wrap_text=True)

                    # descripcion LABEL - data
                    descriptionHeight = 30 + 20 * (len(descripcion) // 80)
                    sheet.row_dimensions[7].height = descriptionHeight
                    sheet['A7'].font = Arial11Bold
                    sheet['B7'].font = Calibri10

                    sheet['A7'].border = thin_border
                    sheet['B7'].border = thin_border
                    sheet['C7'].border = thin_border

                    sheet['A7'].alignment = Alignment(horizontal="center", vertical='center', wrap_text=True)
                    sheet['A7'] = "Descripción de la vulnerabilidad:"
                    sheet.merge_cells('B7:C7')
                    sheet['B7'].alignment = Alignment(horizontal="left", vertical='center', wrap_text=True)
                    sheet['B7'] = descripcion

                    # DETALLES DE LA PRUEBA LABEL
                    sheet.merge_cells('A8:C8')
                    sheet.row_dimensions[8].height = 29
                    sheet['A8'].font = Arial11Bold
                    sheet['A8'].border = thin_border
                    sheet['A8'].alignment = Alignment(horizontal="center", vertical='center', wrap_text=True)
                    sheet['A8'] = "DETALLES DE LA PRUEBA"
                    sheet['A8'].fill = greyFill



                    # HOSTS LABEL - DATA
                    hostHeight1 = 20 + 20 * (len(hosts) // 80)

                    countNL = hosts.count("SALTOLINEA")
                    hostHeight2 = (countNL + 1) * 20
                    if hostHeight1 > hostHeight2:
                        hostHeight = hostHeight1
                    else:
                        hostHeight = hostHeight2
                    sheet.row_dimensions[9].height = hostHeight
                    sheet['A9'].font = Arial11Bold
                    sheet['B9'].font = Calibri10

                    sheet['A9'].border = thin_border
                    sheet['B9'].border = thin_border
                    sheet['C9'].border = thin_border

                    sheet['A9'].alignment = Alignment(horizontal="center", vertical='center', wrap_text=True)
                    sheet['A9'] = "Hosts:"

                    sheet.merge_cells('B9:C9')
                    sheet['B9'].alignment = Alignment(horizontal="left", vertical='center', wrap_text=True)
                    hosts = hosts.replace("SALTOLINEA", "\n")
                    sheet['B9'] = hosts

                    # DETALLES DE LA PRUEBA DATA
                    cellHeight = 20 + 20 * (len(detallesTest) // 80)
                    sheet.merge_cells('A10:C10')
                    sheet.row_dimensions[10].height = cellHeight
                    sheet['A10'].font = Arial10

                    sheet['A10'].border = thin_border
                    sheet['B10'].border = thin_border
                    sheet['C10'].border = thin_border

                    sheet['A10'].alignment = Alignment(horizontal="justify", vertical='center', wrap_text=True)
                    detallesTest = detallesTest.replace("SALTOLINEA", "\n")
                    sheet['A10'] = detallesTest

                    # CONTRAMEDIDAS - LABEL
                    sheet.merge_cells('A11:C11')
                    sheet.row_dimensions[11].height = 29
                    sheet['A11'].font = Arial11Bold

                    sheet['A11'].border = thin_border
                    sheet['B11'].border = thin_border
                    sheet['C11'].border = thin_border

                    sheet['A11'].alignment = Alignment(horizontal="center", vertical='center', wrap_text=True)
                    sheet['A11'] = "CONTRAMEDIDAS"
                    sheet['A11'].fill = greyFill

                    # CONTRAMEDIDAS
                    sheet.merge_cells('A12:C12')
                    cellHeight = 20 + 20 * (len(recomendacion) // 80)

                    sheet.row_dimensions[12].height = cellHeight
                    sheet['A12'].font = Arial10

                    sheet['A12'].border = thin_border
                    sheet['B12'].border = thin_border
                    sheet['C12'].border = thin_border
                    sheet['A12'].alignment = Alignment(horizontal="justify", vertical='center', wrap_text=True)
                    recomendacion = recomendacion.replace("SALTOLINEA", "\n")
                    sheet['A12'] = recomendacion

                    # REFERENCIAS - LABEL
                    sheet.merge_cells('A13:C13')
                    sheet.row_dimensions[13].height = 29
                    sheet['A13'].font = Arial11Bold

                    sheet['A13'].border = thin_border
                    sheet['B13'].border = thin_border
                    sheet['C13'].border = thin_border

                    sheet['A13'].alignment = Alignment(horizontal="center", vertical='center', wrap_text=True)
                    sheet['A13'] = "REFERENCIAS"
                    sheet['A13'].fill = greyFill

                    # REFERENCIAS
                    sheet.merge_cells('A14:C14')
                    cellHeight = 20 + 20 * (len(referencias) // 80)
                    sheet.row_dimensions[14].height = cellHeight
                    sheet['A14'].font = Arial10

                    sheet['A14'].border = thin_border
                    sheet['B14'].border = thin_border
                    sheet['C14'].border = thin_border

                    sheet['A14'].alignment = Alignment(horizontal="justify", vertical='center', wrap_text=True)
                    referencias = referencias.replace("SALTOLINEA", "\n")
                    referencias = referencias.replace("TAB", "\t")
                    sheet['A14'] = referencias



                ##########  BISIT #########
                if (empresa == "BISIT"):
                    #globals()['conclusiones_list_' + vector].append(conclusion)
                    #globals()['recomendaciones_list_' + vector].append(recomendacionMatriz)

                    globals()['wb' + empresa].create_sheet()
                    sheet = globals()['wb' + empresa]['Sheet']

                    sheet.column_dimensions['A'].width = 41
                    sheet.column_dimensions['B'].width = 41

                    sheet.title = str(numeroVulnerabilidad)
                    if ("CRÍTICO" in riesgoInforme):
                        sheet['A5'].font = MuyAlto

                    if ("ALTO" in riesgoInforme):
                        sheet['A5'].font = Alto

                    if ("MEDIO" in riesgoInforme):
                        sheet['A5'].font = Moderado

                    if ("BAJO" in riesgoInforme):
                        sheet['A5'].font = Bajo

                    # Infome modelo BISIT
                    # Titulo
                    sheet['A1'].font = Arial11Bold
                    sheet['A1'] = str(nombre)
                    sheet['A1'].border = thin_border
                    sheet['B1'].border = thin_border

                    # Vulnerabilidad LABEL
                    sheet.merge_cells('A2:B2')
                    sheet['A2'].alignment = Alignment(horizontal="center", vertical='center', wrap_text=True)
                    sheet['A2'].border = thin_border
                    sheet['B2'].border = thin_border
                    sheet['A2'] = "VULNERABILIDAD"
                    sheet['A2'].fill = brownFill
                    sheet['A2'].font = Arial11Bold

                    # Vulnerabilidad
                    sheet.merge_cells('A3:B3')
                    sheet['A3'].alignment = Alignment(horizontal="justify", vertical='center', wrap_text=True)
                    sheet['A3'].border = thin_border
                    sheet['B3'].border = thin_border
                    sheet['A3'] = descripcion
                    # descriptionHeight = 20 + 20 * (len(descripcion) // 80)
                    # sheet.row_dimensions[3].height = descriptionHeight

                    # FACTOR DE RIESGO - LABEL
                    sheet.merge_cells('A4:B4')
                    sheet['A4'].alignment = Alignment(horizontal="center", vertical='center', wrap_text=True)
                    sheet['A4'].border = thin_border
                    sheet['B4'].border = thin_border
                    sheet['A4'] = "FACTOR DE RIESGO"
                    sheet['A4'].fill = brownFill
                    sheet['A4'].font = Arial11Bold

                    # FACTOR DE RIESGO
                    sheet['A5'].alignment = Alignment(horizontal="left", vertical='center', wrap_text=True)
                    sheet['B5'].alignment = Alignment(horizontal="center", vertical='center', wrap_text=True)
                    sheet['A5'].border = thin_border
                    sheet['B5'].border = thin_border
                    sheet['A5'] = riesgoInforme.upper()
                    sheet['B5'] = "CVSS Base Score: " + str(score)

                    # PARA CONOCER MAS ACERCA DE LA VULNERABILIDAD, CONSULTE EN INTERNET:
                    sheet.merge_cells('A6:B6')
                    sheet['A6'].alignment = Alignment(horizontal="center", vertical='center', wrap_text=True)
                    sheet['A6'].border = thin_border
                    sheet['B6'].border = thin_border
                    sheet['A6'] = "PARA CONOCER MAS ACERCA DE LA VULNERABILIDAD, CONSULTE EN INTERNET:"
                    sheet['A6'].fill = brownFill
                    sheet['A6'].font = Arial10Bold

                    # REFERENCIAS
                    sheet.merge_cells('A7:B7')
                    print(f"referencias ")
                    sheet['A7'].alignment = Alignment(horizontal="left", vertical='center', wrap_text=True)
                    sheet['A7'].border = thin_border
                    sheet['B7'].border = thin_border
                    cellHeight = 20 + 20 * (len(referencias) // 80)
                    sheet.row_dimensions[7].height = cellHeight
                    sheet['A7'].font = Arial10
                    referencias = referencias.replace("SALTOLINEA", "\n")
                    referencias = referencias.replace("TAB", "\t")
                    sheet['A7'] = referencias

                    # EXPLOTACION - LABEL
                    sheet.merge_cells('A8:B8')
                    sheet['A8'].alignment = Alignment(horizontal="center", vertical='center', wrap_text=True)
                    sheet['A8'].border = thin_border
                    sheet['B8'].border = thin_border
                    sheet['A8'] = "EXPLOTACION"
                    sheet['A8'].fill = brownFill
                    sheet['A8'].font = Arial11Bold

                    # EXPLOTACION
                    sheet.merge_cells('A9:B9')
                    sheet['A9'].alignment = Alignment(horizontal="justify", vertical='center', wrap_text=True)
                    sheet['A9'].border = thin_border
                    sheet['B9'].border = thin_border
                    cellHeight = 20 + 20 * (len(detallesTest) // 80)
                    sheet.row_dimensions[9].height = cellHeight
                    sheet['A9'].font = Arial10
                    detallesTest = detallesTest.replace("SALTOLINEA", "\n")
                    sheet['A9'] = "POSITIVA," + detallesTest

                    # HOST - LABEL
                    sheet.merge_cells('A10:B10')
                    sheet['A10'].alignment = Alignment(horizontal="center", vertical='center', wrap_text=True)
                    sheet['A10'].border = thin_border
                    sheet['B10'].border = thin_border
                    sheet['A10'] = "HOSTS AFECTADOS"
                    sheet['A10'].fill = brownFill
                    sheet['A10'].font = Arial11Bold

                    # HOST
                    sheet.merge_cells('A11:B11')
                    sheet['A11'].alignment = Alignment(horizontal="left", vertical='center', wrap_text=True)
                    sheet['A11'].border = thin_border
                    sheet['B11'].border = thin_border

                    hostHeight1 = 20 + 20 * (len(hosts) // 80)
                    countNL = hosts.count("SALTOLINEA")
                    hostHeight2 = (countNL + 1) * 20
                    if hostHeight1 > hostHeight2:
                        hostHeight = hostHeight1
                    else:
                        hostHeight = hostHeight2
                    sheet.row_dimensions[11].height = hostHeight
                    sheet['A11'].font = Arial10
                    hosts = hosts.replace("SALTOLINEA", "\n")
                    sheet['A11'] = hosts

                    # CONTRAMEDIDAS - LABEL
                    sheet.merge_cells('A12:B12')
                    sheet['A12'].alignment = Alignment(horizontal="center", vertical='center', wrap_text=True)
                    sheet['A12'].border = thin_border
                    sheet['B12'].border = thin_border
                    sheet['A12'] = "CONTRAMEDIDAS"
                    sheet['A12'].fill = brownFill
                    sheet['A12'].font = Arial11Bold

                    # CONTRAMEDIDAS
                    sheet.merge_cells('A13:B13')
                    sheet['A13'].alignment = Alignment(horizontal="justify", vertical='center', wrap_text=True)
                    sheet['A13'].border = thin_border
                    sheet['B13'].border = thin_border
                    cellHeight = 20 + 20 * (len(recomendacion) // 80)
                    sheet.row_dimensions[13].height = cellHeight
                    sheet['A13'].font = Arial10
                    recomendacion = recomendacion.replace("SALTOLINEA", "\n")
                    recomendacion = recomendacion.replace("TAB", "\t")
                    sheet['A13'] = recomendacion

                numeroVulnerabilidad = numeroVulnerabilidad + 1

        ##### CONCLUSIONES Y RECOMENDACIONES ####
        print("Generando conclusiones y recomendaciones")
        globals()['wb' + empresa].create_sheet()
        sheet = globals()['wb' + empresa]['Sheet']
        sheet.title = 'Conclusiones y recomendaciones'  # Change title.

        #if "BISIT" in empresa:
        if (empresa == "BISIT"):
            sheet.column_dimensions['A'].width = 100
            print ("Reporte BISITT")
            i = 1
            for tupla in matrizRecomendaciones:
                #print (f"tupa {tupla}")
                activo = tupla["activo"]
                conclusion = tupla["conclusion"]
                recomendacionMatriz = tupla["recomendacionMatriz"]
                riesgo = tupla["riesgo"]
                vectorInforme = tupla["vectorInforme"]
                vulnerabilidad = tupla["vulnerabilidad"]
                sheet['A' + str(i)] = conclusion
                sheet['A' + str(i + len(matrizRecomendaciones)+1)] = "Recomendación " + str(i) + ": " + recomendacionMatriz
                i = i + 1

        #if empresa in "Radical":
        if (empresa == "Radical"):
            i = 1
            #pprint.pprint(matrizRecomendaciones)
            sheet.column_dimensions['A'].width = 5
            sheet.column_dimensions['B'].width = 18
            sheet.column_dimensions['C'].width = 23
            sheet.column_dimensions['D'].width = 10
            sheet.column_dimensions['E'].width = 32
            sheet['A' + str(i)] = "Nº"
            sheet['B' + str(i)] = "Activo/Recurso"
            sheet['C' + str(i)] = "Vulnerabilidad identificada"
            sheet['D' + str(i)] = "Nivel de riesgo"
            sheet['E' + str(i)] = "Contramedida "
            sheet['A' + str(i)].font = Calibri12Bold
            sheet['B' + str(i)].font = Calibri12Bold
            sheet['C' + str(i)].font = Calibri12Bold
            sheet['D' + str(i)].font = Calibri12Bold
            sheet['E' + str(i)].font = Calibri12Bold
            sheet['A' + str(i)].alignment = Alignment(horizontal="center", vertical='center', wrap_text=True)
            sheet['B' + str(i)].alignment = Alignment(horizontal="center", vertical='center', wrap_text=True)
            sheet['C' + str(i)].alignment = Alignment(horizontal="center", vertical='center', wrap_text=True)
            sheet['D' + str(i)].alignment = Alignment(horizontal="center", vertical='center', wrap_text=True)
            sheet['E' + str(i)].alignment = Alignment(horizontal="center", vertical='center', wrap_text=True)
            sheet['A' + str(i)].border = thin_border
            sheet['B' + str(i)].border = thin_border
            sheet['C' + str(i)].border = thin_border
            sheet['D' + str(i)].border = thin_border
            sheet['E' + str(i)].border = thin_border



            for tupla in matrizRecomendaciones:
                activo = tupla["activo"]
                conclusion = tupla["conclusion"]
                recomendacionMatriz = tupla["recomendacionMatriz"]
                riesgo = tupla["riesgo"]
                vectorInforme = tupla["vectorInforme"]
                vulnerabilidad = tupla["vulnerabilidad"]
                sheet['A' + str(i+1)] = i
                sheet['B' + str(i+1)] = activo
                sheet['C' + str(i+1)] = vulnerabilidad
                sheet['D' + str(i+1)] = riesgo
                sheet['E' + str(i+1)] = recomendacionMatriz
                sheet['A' + str(i + 1)].alignment = Alignment(horizontal="center", vertical='center',wrap_text=True)
                sheet['B' + str(i + 1)].alignment = Alignment(horizontal="center", vertical='center', wrap_text=True)
                sheet['C' + str(i + 1)].alignment = Alignment(horizontal="center", vertical='center', wrap_text=True)
                sheet['D' + str(i + 1)].alignment = Alignment(horizontal="center", vertical='center', wrap_text=True)
                sheet['E' + str(i + 1)].alignment = Alignment(horizontal="center", vertical='center', wrap_text=True)
                sheet['A' + str(i + 1)].border = thin_border
                sheet['B' + str(i + 1)].border = thin_border
                sheet['C' + str(i + 1)].border = thin_border
                sheet['D' + str(i + 1)].border = thin_border
                sheet['E' + str(i + 1)].border = thin_border

                if ("CRÍTICO" in riesgo):
                    sheet['D' + str(i+1)].fill = criticoFill

                if ("ALTO" in riesgo):
                    sheet['D' + str(i+1)].fill = altoFill

                if ("MEDIO" in riesgo):
                    sheet['D' + str(i+1)].fill = medioFill

                if ("BAJO" in riesgo):
                    sheet['D' + str(i+1)].fill = bajoFill
                i = i + 1
        # escribir archivo
        globals()['wb' + empresa].save(f'{empresa}-{vectorInforme}.xlsx')


###### INFORME EJECUTIVO ####
# crear nueva pestania
wbEjecutivo.create_sheet()
sheet = wbEjecutivo['Sheet']

sheet.title = 'Vulnerabilidades explotadas'  # Change title

sheet.column_dimensions['A'].width = 20
sheet.column_dimensions['B'].width = 25
sheet['A1'] = "Valor"
sheet['A2'] = "Crítico"
sheet['A3'] = "Alto"
sheet['A4'] = "Medio"
sheet['A5'] = "Bajo"
#sheet['A6'] = "Muy bajo"

sheet['A1'].border = thin_border
sheet['A2'].border = thin_border
sheet['A3'].border = thin_border
sheet['A4'].border = thin_border
sheet['A5'].border = thin_border
#sheet['A6'].border = thin_border

sheet['B1'].border = thin_border
sheet['B2'].border = thin_border
sheet['B3'].border = thin_border
sheet['B4'].border = thin_border
sheet['B5'].border = thin_border
#sheet['B6'].border = thin_border

sheet['A1'].alignment = Alignment(horizontal="center", vertical='center', wrap_text=True)
sheet['A2'].alignment = Alignment(horizontal="center", vertical='center', wrap_text=True)
sheet['A3'].alignment = Alignment(horizontal="center", vertical='center', wrap_text=True)
sheet['A4'].alignment = Alignment(horizontal="center", vertical='center', wrap_text=True)
sheet['A5'].alignment = Alignment(horizontal="center", vertical='center', wrap_text=True)
#sheet['A6'].alignment = Alignment(horizontal="center", vertical='center', wrap_text=True)

sheet['B1'].alignment = Alignment(horizontal="center", vertical='center', wrap_text=True)
sheet['B2'].alignment = Alignment(horizontal="center", vertical='center', wrap_text=True)
sheet['B3'].alignment = Alignment(horizontal="center", vertical='center', wrap_text=True)
sheet['B4'].alignment = Alignment(horizontal="center", vertical='center', wrap_text=True)
sheet['B5'].alignment = Alignment(horizontal="center", vertical='center', wrap_text=True)
#sheet['B6'].alignment = Alignment(horizontal="center", vertical='center', wrap_text=True)

sheet['A1'].fill = negroFill
sheet['A1'].font = Arial12BoldWhite
sheet['A2'].fill = criticoFill
sheet['A3'].fill = AltoFill2
sheet['A4'].fill = ModeradoFill
sheet['A5'].fill = bajoFill2
#sheet['A6'].fill = MuyBajoFill

sheet['B1'].fill = negroFill
sheet['B1'].font = Arial12BoldWhite
sheet['B1'] = "Número de riesgos"
sheet['B2'] = total_vul_criticas_INTERNO + total_vul_criticas_EXTERNO
sheet['B3'] = total_vul_altas_INTERNO + total_vul_altas_EXTERNO
sheet['B4'] = total_vul_medias_INTERNO + total_vul_medias_EXTERNO
sheet['B5'] = total_vul_bajas_INTERNO + total_vul_bajas_EXTERNO
#sheet['B6'] = 0


chart = BarChart()
chart.type = "col"
chart.style = 10
chart.shape = 4
chart.title = "Vulnerabilidades explotadas"


#chart.y_axis.title = 'Test number'
#chart.x_axis.title = 'Vulnerabilidades explotadas'

# create data for plotting
labels = Reference( sheet, min_col=1, min_row=2, max_row=5)
data = Reference( sheet, min_col=2, min_row=2, max_row=5)

# adding data to the Doughnut chart object
chart.add_data(data, titles_from_data=False)
chart.set_categories(labels)

chart.dataLabels = DataLabelList()
chart.dataLabels.showPercent = True
chart.dataLabels.showVal = False
chart.dataLabels.showLegendKey = True
chart.dataLabels.showCatName = False

# try to set color blue (0000FF) for the 2nd wedge (idx=1) in the series
series = chart.series[0]
pt = DataPoint(idx=0)
pt.graphicalProperties.solidFill = "FF0000"
series.dPt.append(pt)

pt = DataPoint(idx=1)
pt.graphicalProperties.solidFill = "FFC000"
series.dPt.append(pt)

pt = DataPoint(idx=2)
pt.graphicalProperties.solidFill = "FFFF00"
series.dPt.append(pt)

pt = DataPoint(idx=3)
pt.graphicalProperties.solidFill = "92D050"
series.dPt.append(pt)

#pt = DataPoint(idx=4)
#pt.graphicalProperties.solidFill = "95B3D7"
#series.dPt.append(pt)
#adicionar la grafica a la hoja de calculo
sheet.add_chart(chart, 'C5')




###### Vulnerabilidades externas vs internas ####
        # crear nueva pestania
wbEjecutivo.create_sheet()
sheet = wbEjecutivo['Sheet']

sheet.title = 'externas vs internas'  # Change title

sheet.column_dimensions['A'].width = 20
sheet.column_dimensions['B'].width = 25
sheet['A1'] = "Vector"
sheet['A2'] = "Externas"
sheet['A3'] = "Internas"


sheet['A1'].border = thin_border
sheet['A2'].border = thin_border
sheet['A3'].border = thin_border


sheet['B1'].border = thin_border
sheet['B2'].border = thin_border
sheet['B3'].border = thin_border


sheet['A1'].alignment = Alignment(horizontal="center", vertical='center', wrap_text=True)
sheet['A2'].alignment = Alignment(horizontal="center", vertical='center', wrap_text=True)
sheet['A3'].alignment = Alignment(horizontal="center", vertical='center', wrap_text=True)


sheet['B1'].alignment = Alignment(horizontal="center", vertical='center', wrap_text=True)
sheet['B2'].alignment = Alignment(horizontal="center", vertical='center', wrap_text=True)
sheet['B3'].alignment = Alignment(horizontal="center", vertical='center', wrap_text=True)



sheet['A1'].fill = negroFill
sheet['A1'].font = Arial12BoldWhite
sheet['A1'] = "Vector"
sheet['B1'].fill = negroFill
sheet['B1'].font = Arial12BoldWhite
sheet['B1'] = "Vulnerabilidades"

sheet['B2'] = total_vul_criticas_EXTERNO + total_vul_altas_EXTERNO + total_vul_medias_EXTERNO + total_vul_bajas_EXTERNO
sheet['B3'] = total_vul_criticas_INTERNO + total_vul_altas_INTERNO +total_vul_medias_INTERNO + total_vul_bajas_INTERNO

chart = PieChart3D()
chart.title = "Vulnerabilidades por vector"
# create data for plotting
labels = Reference( sheet, min_col=1, min_row=2, max_row=3)
data = Reference( sheet, min_col=2, min_row=2, max_row=3)

# adding data to the Doughnut chart object
chart.add_data(data, titles_from_data=False)
chart.set_categories(labels)

chart.dataLabels = DataLabelList()
chart.dataLabels.showPercent = True
chart.dataLabels.showVal = False
chart.dataLabels.showLegendKey = False
chart.dataLabels.showCatName = False

# set style of the chart
chart.style = 26

# try to set color blue (0000FF) for the 2nd wedge (idx=1) in the series
series = chart.series[0]
pt = DataPoint(idx=0)
pt.graphicalProperties.solidFill = "B74C49"
series.dPt.append(pt)

pt = DataPoint(idx=1)
pt.graphicalProperties.solidFill = "4B7BB4"
series.dPt.append(pt)

#adicionar la grafica a la hoja de calculo
sheet.add_chart(chart, 'C5')



###### Vulnerabilidades por tipo de activo ####
        # crear nueva pestania
wbEjecutivo.create_sheet()
sheet = wbEjecutivo['Sheet']

sheet.title = 'activos de información'  # Change title

sheet.column_dimensions['A'].width = 20
sheet.column_dimensions['B'].width = 25

sheet['A2'] = "Servidores"
sheet['A3'] = "Base de datos"
sheet['A4'] = "VoIP"
sheet['A5'] = "Sistemas de vigilancia"
sheet['A6'] = "Dispositivos de red"
sheet['A7'] = "Personal"
sheet['A8'] = "Aplicación web"
sheet['A9'] = "Estaciones de trabajo"
sheet['A10'] = "Otros"

sheet['A1'].border = thin_border
sheet['A2'].border = thin_border
sheet['A3'].border = thin_border
sheet['A4'].border = thin_border
sheet['A5'].border = thin_border
sheet['A6'].border = thin_border
sheet['A7'].border = thin_border
sheet['A8'].border = thin_border
sheet['A9'].border = thin_border
sheet['A10'].border = thin_border


sheet['B1'].border = thin_border
sheet['B2'].border = thin_border
sheet['B3'].border = thin_border
sheet['B4'].border = thin_border
sheet['B5'].border = thin_border
sheet['B6'].border = thin_border
sheet['B7'].border = thin_border
sheet['B8'].border = thin_border
sheet['B9'].border = thin_border
sheet['B10'].border = thin_border


sheet['A1'].alignment = Alignment(horizontal="center", vertical='center', wrap_text=True)
sheet['A2'].alignment = Alignment(horizontal="center", vertical='center', wrap_text=True)
sheet['A3'].alignment = Alignment(horizontal="center", vertical='center', wrap_text=True)
sheet['A4'].alignment = Alignment(horizontal="center", vertical='center', wrap_text=True)
sheet['A5'].alignment = Alignment(horizontal="center", vertical='center', wrap_text=True)
sheet['A6'].alignment = Alignment(horizontal="center", vertical='center', wrap_text=True)
sheet['A7'].alignment = Alignment(horizontal="center", vertical='center', wrap_text=True)
sheet['A8'].alignment = Alignment(horizontal="center", vertical='center', wrap_text=True)
sheet['A9'].alignment = Alignment(horizontal="center", vertical='center', wrap_text=True)
sheet['A10'].alignment = Alignment(horizontal="center", vertical='center', wrap_text=True)


sheet['B1'].alignment = Alignment(horizontal="center", vertical='center', wrap_text=True)
sheet['B2'].alignment = Alignment(horizontal="center", vertical='center', wrap_text=True)
sheet['B3'].alignment = Alignment(horizontal="center", vertical='center', wrap_text=True)
sheet['B4'].alignment = Alignment(horizontal="center", vertical='center', wrap_text=True)
sheet['B5'].alignment = Alignment(horizontal="center", vertical='center', wrap_text=True)
sheet['B6'].alignment = Alignment(horizontal="center", vertical='center', wrap_text=True)
sheet['B7'].alignment = Alignment(horizontal="center", vertical='center', wrap_text=True)
sheet['B8'].alignment = Alignment(horizontal="center", vertical='center', wrap_text=True)
sheet['B9'].alignment = Alignment(horizontal="center", vertical='center', wrap_text=True)
sheet['B10'].alignment = Alignment(horizontal="center", vertical='center', wrap_text=True)


## Cabecera
sheet['A1'].fill = negroFill
sheet['A1'].font = Arial12BoldWhite
sheet['A1'] = "Tipo de activo"
sheet['B1'].fill = negroFill
sheet['B1'].font = Arial12BoldWhite
sheet['B1'] = "Vulnerabilidades"

#datos
sheet['B2'] = servidores
sheet['B3'] = baseDatos
sheet['B4'] = telefoniaIP
sheet['B5'] = sistemaVigilancia
sheet['B6'] = dispositivosRed
sheet['B7'] = personal
sheet['B8'] = aplicacionWeb
sheet['B9'] = estacionesTrabajo
sheet['B10'] = otros


chart = PieChart()
chart.title = "Vulnerabilidades por activos de información"
#chart.style = 26

# create data for plotting
labels = Reference( sheet, min_col=1, min_row=2, max_row=10)
data = Reference( sheet, min_col=2, min_row=2, max_row=10)

# adding data to the Doughnut chart object
chart.add_data(data, titles_from_data=False)
chart.set_categories(labels)

# agregar espacio
slice = DataPoint(idx=0, explosion=20)
chart.series[0].data_points = [slice]
#chart.series[1].data_points = [slice]
#chart.series[2].data_points = [slice]

chart.dataLabels = DataLabelList()
chart.dataLabels.showPercent = True
chart.dataLabels.showVal = False
chart.dataLabels.showLegendKey = True
chart.dataLabels.showCatName = True

#adicionar la grafica a la hoja de calculo
sheet.add_chart(chart, 'C5')



### Total pruebas vs explotadas

# crear nueva pestania
wbEjecutivo.create_sheet()
sheet = wbEjecutivo['Sheet']

sheet.title = 'Total pruebas'  # Change title

sheet.column_dimensions['A'].width = 20
sheet.column_dimensions['B'].width = 25
sheet['A1'] = "Total pruebas específicas"
sheet['A2'] = "Vulnerabilidades identificadas"

sheet['A1'].border = thin_border
sheet['A2'].border = thin_border


sheet['B1'].border = thin_border
sheet['B2'].border = thin_border

sheet['A1'].alignment = Alignment(horizontal="center", vertical='center', wrap_text=True)
sheet['A2'].alignment = Alignment(horizontal="center", vertical='center', wrap_text=True)


sheet['B1'].alignment = Alignment(horizontal="center", vertical='center', wrap_text=True)
sheet['B2'].alignment = Alignment(horizontal="center", vertical='center', wrap_text=True)



sheet['B1'] = totalPruebas
sheet['B2'] = total_vul_criticas_INTERNO + total_vul_criticas_EXTERNO + total_vul_altas_INTERNO + total_vul_altas_EXTERNO +total_vul_medias_INTERNO + total_vul_medias_EXTERNO + total_vul_bajas_INTERNO + total_vul_bajas_EXTERNO



chart = BarChart()
chart.type = "bar"
chart.style = 10
chart.shape = 4
chart.title = "Vulnerabilidades explotadas"


#chart.y_axis.title = 'Test number'
#chart.x_axis.title = 'Vulnerabilidades explotadas'

# create data for plotting
labels = Reference( sheet, min_col=1, min_row=1, max_row=2)
data = Reference( sheet, min_col=2, min_row=1, max_row=2)

# adding data to the Doughnut chart object
chart.add_data(data, titles_from_data=False)
chart.set_categories(labels)

chart.dataLabels = DataLabelList()
chart.dataLabels.showPercent = True
chart.dataLabels.showVal = True
chart.dataLabels.showLegendKey = True
chart.dataLabels.showCatName = False

# try to set color blue (0000FF) for the 2nd wedge (idx=1) in the series
series = chart.series[0]
pt = DataPoint(idx=0)
pt.graphicalProperties.solidFill = "007BD4"
series.dPt.append(pt)

pt = DataPoint(idx=1)
pt.graphicalProperties.solidFill = "FB770B"
series.dPt.append(pt)

#adicionar la grafica a la hoja de calculo
sheet.add_chart(chart, 'C5')




###### Vulnerabilidades por tipo de vulnerabilidad ####
        # crear nueva pestania
wbEjecutivo.create_sheet()
sheet = wbEjecutivo['Sheet']

sheet.title = 'Tipo vulnerabilidad'  # Change title

sheet.column_dimensions['A'].width = 20
sheet.column_dimensions['B'].width = 25

sheet['A2'] = "Password débil"
sheet['A3'] = "Falta parches de seguridad"
sheet['A4'] = "Errores de configuración"

sheet['A1'].border = thin_border
sheet['A2'].border = thin_border
sheet['A3'].border = thin_border
sheet['A4'].border = thin_border


sheet['B1'].border = thin_border
sheet['B2'].border = thin_border
sheet['B3'].border = thin_border
sheet['B4'].border = thin_border


sheet['A1'].alignment = Alignment(horizontal="center", vertical='center', wrap_text=True)
sheet['A2'].alignment = Alignment(horizontal="center", vertical='center', wrap_text=True)
sheet['A3'].alignment = Alignment(horizontal="center", vertical='center', wrap_text=True)
sheet['A4'].alignment = Alignment(horizontal="center", vertical='center', wrap_text=True)


sheet['B1'].alignment = Alignment(horizontal="center", vertical='center', wrap_text=True)
sheet['B2'].alignment = Alignment(horizontal="center", vertical='center', wrap_text=True)
sheet['B3'].alignment = Alignment(horizontal="center", vertical='center', wrap_text=True)
sheet['B4'].alignment = Alignment(horizontal="center", vertical='center', wrap_text=True)


## Cabecera
sheet['A1'].fill = negroFill
sheet['A1'].font = Arial12BoldWhite
sheet['A1'] = "Tipo"
sheet['B1'].fill = negroFill
sheet['B1'].font = Arial12BoldWhite
sheet['B1'] = "Vulnerabilidades"

#datos
sheet['B2'] = passwordDebil
sheet['B3'] = faltaParches
sheet['B4'] = errorConfiguracion


chart = PieChart3D()
chart.title = "Vulnerabilidades por tipo"
#chart.style = 26

# create data for plotting
labels = Reference( sheet, min_col=1, min_row=2, max_row=4)
data = Reference( sheet, min_col=2, min_row=2, max_row=4)

# adding data to the Doughnut chart object
chart.add_data(data, titles_from_data=False)
chart.set_categories(labels)


chart.dataLabels = DataLabelList()
chart.dataLabels.showPercent = True
chart.dataLabels.showVal = False
chart.dataLabels.showLegendKey = False
chart.dataLabels.showCatName = False


#adicionar la grafica a la hoja de calculo
sheet.add_chart(chart, 'C5')

wbEjecutivo.save(f'informe-ejecutivo.xlsx')

