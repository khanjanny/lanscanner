#!/usr/bin/python3

import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.chart import PieChart3D, Reference 
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.chart.marker import DataPoint
from openpyxl.chart.label import DataLabelList
import math
import itertools
import csv

# Fuentes
MuyAlto = Font(name='Arial', size=12, bold=True, color='FC611C')
Alto = Font(name='Arial', size=12, bold=True, color='FFC000')
Moderado = Font(name='Arial', size=12, bold=True, color='FFFF00')
Bajo = Font(name='Arial', size=12, bold=True, color='95B3D7')
MuyBajo = Font(name='Arial', size=12, bold=True, color='92D050')

italic24Font = Font(size=24, italic=True, bold=True)
Arial10 = Font(name='Arial', size=10)
Arial11Bold = Font(name='Arial', size=11, bold=True)
Arial12BoldWhite = Font(name='Arial', size=12, bold=True, color='FFFFFF')
Arial10Bold = Font(name='Arial', size=10, bold=True)
Calibri10 = Font(name='Calibri', size=10)

Impact22 = Font(name='Impact', size=22)
Impact36 = Font(name='Impact', size=36)

# Borde
thin_border = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))

# Rellenos
skyBlueFill = PatternFill(start_color='bdd7ee',
                          end_color='bdd7ee',
                          fill_type='solid')

greyFill = PatternFill(start_color='d9d9d9',
                       end_color='d9d9d9',
                       fill_type='solid')


#### ISB #####
criticoFill = PatternFill(start_color='FF0000',
                          end_color='FF0000',
                          fill_type='solid')

altoFill = PatternFill(start_color='FFC000',
                       end_color='FFC000',
                       fill_type='solid')

medioFill = PatternFill(start_color='FFFF00',
                        end_color='FFFF00',
                        fill_type='solid')

bajoFill = PatternFill(start_color='00B050',
                       end_color='00B050',
                       fill_type='solid')

brownFill = PatternFill(start_color='EEECE1',
                        end_color='EEECE1',
                        fill_type='solid')
                        

#### BISIT #####
MuyAltoFill = PatternFill(start_color='FC611C', end_color='FC611C', fill_type='solid')
AltoFill2 = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
ModeradoFill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
bajoFill2 = PatternFill(start_color='95B3D7', end_color='95B3D7', fill_type='solid')
MuyBajoFill = PatternFill(start_color='92D050', end_color='92D050', fill_type='solid')    
negroFill = PatternFill(start_color='17150D', end_color='17150D', fill_type='solid')    

                  
                 
##############

vulnerabilidadesDF = pd.read_csv("datos.csv", sep='|', encoding='utf-8')

wbBISIT = openpyxl.Workbook()  # Create a blank workbook.
sheet = wbBISIT['Sheet']
sheet.title = 'stats'  # Change title.




with open('reporte-resumen.csv', 'r') as reporteResumen:
    VulCriticas =  int(next(itertools.islice(csv.reader(reporteResumen), 18, None))[0].split(';')[1]) #dividir por salto de linea    
    VulAltas = int(next(itertools.islice(csv.reader(reporteResumen), 0, None))[0].split(';')[1])
    VulMedias = int(next(itertools.islice(csv.reader(reporteResumen), 0, None))[0].split(';')[1])
    VulBajas = int(next(itertools.islice(csv.reader(reporteResumen), 0, None))[0].split(';')[1])

sheet.column_dimensions['A'].width = 20
sheet.column_dimensions['B'].width = 25
sheet['A1'] = "Valor"
sheet['A2'] = "Muy alto"
sheet['A3'] = "Alto"
sheet['A4'] = "Medio"
sheet['A5'] = "Bajo"
sheet['A6'] = "Muy bajo"


sheet['A1'].border = thin_border
sheet['A2'].border = thin_border
sheet['A3'].border = thin_border
sheet['A4'].border = thin_border
sheet['A5'].border = thin_border
sheet['A6'].border = thin_border

sheet['B1'].border = thin_border
sheet['B2'].border = thin_border
sheet['B3'].border = thin_border
sheet['B4'].border = thin_border
sheet['B5'].border = thin_border
sheet['B6'].border = thin_border


sheet['A1'].alignment = Alignment(horizontal="center", vertical='center', wrap_text=True)
sheet['A2'].alignment = Alignment(horizontal="center", vertical='center', wrap_text=True)
sheet['A3'].alignment = Alignment(horizontal="center", vertical='center', wrap_text=True)
sheet['A4'].alignment = Alignment(horizontal="center", vertical='center', wrap_text=True)
sheet['A5'].alignment = Alignment(horizontal="center", vertical='center', wrap_text=True)
sheet['A6'].alignment = Alignment(horizontal="center", vertical='center', wrap_text=True)

sheet['B1'].alignment = Alignment(horizontal="center", vertical='center', wrap_text=True)
sheet['B2'].alignment = Alignment(horizontal="center", vertical='center', wrap_text=True)
sheet['B3'].alignment = Alignment(horizontal="center", vertical='center', wrap_text=True)
sheet['B4'].alignment = Alignment(horizontal="center", vertical='center', wrap_text=True)
sheet['B5'].alignment = Alignment(horizontal="center", vertical='center', wrap_text=True)
sheet['B6'].alignment = Alignment(horizontal="center", vertical='center', wrap_text=True)

sheet['A1'].fill = negroFill
sheet['A1'].font = Arial12BoldWhite
sheet['A2'].fill = MuyAltoFill
sheet['A3'].fill = AltoFill2
sheet['A4'].fill = ModeradoFill
sheet['A5'].fill = bajoFill2
sheet['A6'].fill = MuyBajoFill


sheet['B1'].fill = negroFill
sheet['B1'].font = Arial12BoldWhite
sheet['B1'] = "Número de riesgos"
sheet['B2'] = VulCriticas
sheet['B3'] = VulAltas
sheet['B4'] = VulMedias
sheet['B5'] = VulBajas
sheet['B6'] = 0

chart = PieChart3D() 
  
# create data for plotting 
labels = Reference(sheet, min_col = 1, min_row = 2, max_row = 6) 
data = Reference(sheet, min_col = 2, min_row = 2, max_row = 6) 

# adding data to the Doughnut chart object 
chart.add_data(data, titles_from_data = False) 
chart.set_categories(labels) 

chart.dataLabels = DataLabelList()
chart.dataLabels.showPercent = True
chart.dataLabels.showVal  = False
chart.dataLabels.showLegendKey = False
chart.dataLabels.showCatName = False
  
# set style of the chart 
chart.style = 26

# try to set color blue (0000FF) for the 2nd wedge (idx=1) in the series
series = chart.series[0]
pt = DataPoint(idx=0)
pt.graphicalProperties.solidFill = "FC611C"
series.dPt.append(pt)

pt = DataPoint(idx=1)
pt.graphicalProperties.solidFill = "FFC000"
series.dPt.append(pt)

pt = DataPoint(idx=2)
pt.graphicalProperties.solidFill = "FFFF00"
series.dPt.append(pt)

pt = DataPoint(idx=3)
pt.graphicalProperties.solidFill = "95B3D7"
series.dPt.append(pt)

pt = DataPoint(idx=4)
pt.graphicalProperties.solidFill = "92D050"
series.dPt.append(pt)


sheet.add_chart(chart, 'C5')

print("Generando reporte BISIT")

for index, row in vulnerabilidadesDF.iterrows():
    contador = row['contador']
    nombre = row['nombre']
    codVul = row['codVul']
    cvss = row['cvss']
    descripcion = row['descripcion']
    agente_amenaza = row['agente_amenaza']
    impacto_tecnico = row['impacto_tecnico']
    impacto_negocio = row['impacto_negocio']
    probabilidad = row['probabilidad']
    impacto = row['impacto']
    riesgoInforme = str(row['riesgoInforme'])
    detallesTest = row['detallesTest']
    recomendacion = row['recomendacion']
    referencias = row['referencias']
    hosts = row['hosts']
    print ("cvss "+ str(cvss))
    if math.isnan(cvss):
        cvss = "N/A"


    print(nombre)

    wbBISIT.create_sheet()
    sheet = wbBISIT['Sheet']

    sheet.column_dimensions['A'].width = 41
    sheet.column_dimensions['B'].width = 41

    sheet.title = str(contador)
    if ("CRÍTICO" in riesgoInforme):
        sheet['A5'].font = MuyAlto

    if ("ALTO" in riesgoInforme):
        sheet['A5'].font = Alto

    if ("MEDIO" in riesgoInforme):
        sheet['A5'].font = Moderado

    if ("BAJO" in riesgoInforme):
        sheet['A5'].font = Bajo

    # Titulo
	
    sheet['A1'].font = Arial11Bold
    sheet['A1'] = str(nombre)
    sheet['A1'].border = thin_border
    sheet['B1'].border = thin_border

    # Vulnerabilidad LABEL
    sheet.merge_cells('A2:B2')
    sheet['A2'].alignment = Alignment(horizontal="center", vertical='center', wrap_text=True)
    sheet['A2'].border = thin_border
    sheet['B2'].border = thin_border
    sheet['A2'] = "VULNERABILIDAD"
    sheet['A2'].fill = brownFill
    sheet['A2'].font = Arial11Bold

    # Vulnerabilidad
    sheet.merge_cells('A3:B3')
    sheet['A3'].alignment = Alignment(horizontal="justify", vertical='center', wrap_text=True)
    sheet['A3'].border = thin_border
    sheet['B3'].border = thin_border
    sheet['A3'] = descripcion
    descriptionHeight = 20 + 20 * (len(descripcion) // 80)
    sheet.row_dimensions[3].height = descriptionHeight



    # FACTOR DE RIESGO - LABEL
    sheet.merge_cells('A4:B4')
    sheet['A4'].alignment = Alignment(horizontal="center", vertical='center', wrap_text=True)
    sheet['A4'].border = thin_border
    sheet['B4'].border = thin_border
    sheet['A4'] = "FACTOR DE RIESGO"
    sheet['A4'].fill = brownFill
    sheet['A4'].font = Arial11Bold

    # FACTOR DE RIESGO
    sheet['A5'].alignment = Alignment(horizontal="left", vertical='center', wrap_text=True)
    sheet['B5'].alignment = Alignment(horizontal="center", vertical='center', wrap_text=True)
    sheet['A5'].border = thin_border
    sheet['B5'].border = thin_border
    sheet['A5'] = riesgoInforme.upper()
    sheet['B5'] = "CVSS Base Score: "+str(cvss)

   # PARA CONOCER MAS ACERCA DE LA VULNERABILIDAD, CONSULTE EN INTERNET:
    sheet.merge_cells('A6:B6')
    sheet['A6'].alignment = Alignment(horizontal="center", vertical='center', wrap_text=True)
    sheet['A6'].border = thin_border
    sheet['B6'].border = thin_border
    sheet['A6'] = "PARA CONOCER MAS ACERCA DE LA VULNERABILIDAD, CONSULTE EN INTERNET:"
    sheet['A6'].fill = brownFill
    sheet['A6'].font = Arial10Bold

    # REFERENCIAS
    sheet.merge_cells('A7:B7')
    sheet['A7'].alignment = Alignment(horizontal="left", vertical='center', wrap_text=True)
    sheet['A7'].border = thin_border
    sheet['B7'].border = thin_border
    cellHeight = 20 + 20 * (len(referencias) // 80)
    sheet.row_dimensions[7].height = cellHeight
    sheet['A7'].font = Arial10
    referencias = referencias.replace("SALTOLINEA", "\n")
    referencias = referencias.replace("TAB", "\t")
    sheet['A7'] = referencias

    # EXPLOTACION - LABEL
    sheet.merge_cells('A8:B8')
    sheet['A8'].alignment = Alignment(horizontal="center", vertical='center', wrap_text=True)
    sheet['A8'].border = thin_border
    sheet['B8'].border = thin_border
    sheet['A8'] = "EXPLOTACION"
    sheet['A8'].fill = brownFill
    sheet['A8'].font = Arial11Bold

    # EXPLOTACION
    sheet.merge_cells('A9:B9')
    sheet['A9'].alignment = Alignment(horizontal="justify", vertical='center', wrap_text=True)
    sheet['A9'].border = thin_border
    sheet['B9'].border = thin_border
    cellHeight = 20 + 20 * (len(detallesTest) // 80)
    sheet.row_dimensions[9].height = cellHeight
    sheet['A9'].font = Arial10
    sheet['A9'] = "POSITIVA,"+detallesTest

    # HOST - LABEL
    sheet.merge_cells('A10:B10')
    sheet['A10'].alignment = Alignment(horizontal="center", vertical='center', wrap_text=True)
    sheet['A10'].border = thin_border
    sheet['B10'].border = thin_border
    sheet['A10'] = "HOSTS AFECTADOS"
    sheet['A10'].fill = brownFill
    sheet['A10'].font = Arial11Bold

    # HOST
    sheet.merge_cells('A11:B11')
    sheet['A11'].alignment = Alignment(horizontal="left", vertical='center', wrap_text=True)
    sheet['A11'].border = thin_border
    sheet['B11'].border = thin_border

    hostHeight1 = 20 + 20 * (len(hosts) // 80)
    countNL = hosts.count("SALTOLINEA")
    hostHeight2 = (countNL + 1) * 20
    if hostHeight1 > hostHeight2:
        hostHeight = hostHeight1
    else:
        hostHeight = hostHeight2
    sheet.row_dimensions[11].height = hostHeight
    sheet['A11'].font = Arial10
    hosts = hosts.replace("SALTOLINEA", "\n")
    sheet['A11'] = hosts


    # CONTRAMEDIDAS - LABEL
    sheet.merge_cells('A12:B12')
    sheet['A12'].alignment = Alignment(horizontal="center", vertical='center', wrap_text=True)
    sheet['A12'].border = thin_border
    sheet['B12'].border = thin_border
    sheet['A12'] = "CONTRAMEDIDAS"
    sheet['A12'].fill = brownFill
    sheet['A12'].font = Arial11Bold

    # CONTRAMEDIDAS
    sheet.merge_cells('A13:B13')
    sheet['A13'].alignment = Alignment(horizontal="justify", vertical='center', wrap_text=True)
    sheet['A13'].border = thin_border
    sheet['B13'].border = thin_border
    cellHeight = 20 + 20 * (len(recomendacion) // 80)
    sheet.row_dimensions[13].height = cellHeight
    sheet['A13'].font = Arial10
    recomendacion = recomendacion.replace("SALTOLINEA", "\n")
    recomendacion = recomendacion.replace("TAB", "\t")
    sheet['A13'] = recomendacion

wbBISIT.save('informeBISIT.xlsx')



print("\n")
print("Generando reporte ISB")

wbISB = openpyxl.Workbook()  # Create a blank workbook.
sheet = wbISB['Sheet']
sheet.title = 'index'  # Change title.

for index, row in vulnerabilidadesDF.iterrows():
    contador = row['contador']
    nombre = row['nombre']
    codVul = row['codVul']
    cvss = row['cvss']
    descripcion = row['descripcion']
    agente_amenaza = row['agente_amenaza']
    impacto_tecnico = row['impacto_tecnico']
    impacto_negocio = row['impacto_negocio']
    probabilidad = row['probabilidad']
    impacto = row['impacto']
    riesgoInforme = str(row['riesgoInforme'])
    detallesTest = row['detallesTest']
    recomendacion = row['recomendacion']
    referencias = row['referencias']
    hosts = row['hosts']

    print(nombre)

    wbISB.create_sheet()
    sheet = wbISB['Sheet']

    sheet.column_dimensions['A'].width = 22
    sheet.column_dimensions['B'].width = 33
    sheet.column_dimensions['C'].width = 27

    sheet.title = str(contador)
    if ("CRÍTICO" in riesgoInforme):
        sheet['A1'].fill = criticoFill
        sheet['C5'].fill = criticoFill

    if ("ALTO" in riesgoInforme):
        sheet['A1'].fill = altoFill
        sheet['C5'].fill = altoFill

    if ("MEDIO" in riesgoInforme):
        sheet['A1'].fill = medioFill
        sheet['C5'].fill = medioFill

    if ("BAJO" in riesgoInforme):
        sheet['A1'].fill = bajoFill
        sheet['C5'].fill = bajoFill

    # Titulo
    sheet.merge_cells('B1:C1')
    sheet.row_dimensions[1].height = 59
    sheet['A1'].font = Impact36
    sheet['B1'].font = Impact22
    sheet['A1'].border = thin_border
    sheet['B1'].border = thin_border
    sheet['C1'].border = thin_border

    sheet['A1'].alignment = Alignment(horizontal="center", vertical='center')
    sheet['B1'].alignment = Alignment(horizontal="center", vertical='center', wrap_text=True)

    sheet['A1'] = str(contador)
    sheet['B1'] = str(nombre).upper()

    # grafico
    sheet.row_dimensions[2].height = 76.5
    img = openpyxl.drawing.image.Image('/usr/share/lanscanner/image.png')

    sheet.add_image(img, "D2")
    sheet['A2'].border = thin_border
    sheet['B2'].border = thin_border
    sheet['C2'].border = thin_border
    sheet.merge_cells('A2:C2')

    # texto del grafico
    sheet.row_dimensions[3].height = 53.2
    sheet['A3'].font = Arial10
    sheet['B3'].font = Arial10
    sheet['C3'].font = Arial10

    sheet['A3'].alignment = Alignment(horizontal="center", vertical='center', wrap_text=True)
    sheet['B3'].alignment = Alignment(horizontal="center", vertical='center', wrap_text=True)
    sheet['C3'].alignment = Alignment(horizontal="center", vertical='center', wrap_text=True)

    sheet['A3'].border = thin_border
    sheet['B3'].border = thin_border
    sheet['C3'].border = thin_border

    sheet['A3'] = str(agente_amenaza)
    sheet['B3'] = str(impacto_tecnico)
    sheet['C3'] = str(impacto_negocio)

    # ANALISIS DE RIESGO LABEL
    sheet.merge_cells('A4:C4')
    sheet.row_dimensions[4].height = 29
    sheet['A4'].font = Arial11Bold

    sheet['A4'].border = thin_border
    sheet['B4'].border = thin_border
    sheet['C4'].border = thin_border

    sheet['A4'].alignment = Alignment(horizontal="center", vertical='center', wrap_text=True)
    sheet['A4'] = "ANALISIS DE RIESGO"
    sheet['A4'].fill = greyFill

    # Riesgos
    sheet.row_dimensions[5].height = 34.5
    sheet['A5'].font = Arial10
    sheet['B5'].font = Arial10
    sheet['C5'].font = Arial11Bold

    sheet['A5'].alignment = Alignment(horizontal="center", vertical='center', wrap_text=True)
    sheet['B5'].alignment = Alignment(horizontal="center", vertical='center', wrap_text=True)
    sheet['C5'].alignment = Alignment(horizontal="center", vertical='center', wrap_text=True)

    sheet['A5'].border = thin_border
    sheet['B5'].border = thin_border
    sheet['C5'].border = thin_border

    sheet['A5'] = "PROBABILIDAD: " + str(probabilidad)
    sheet['B5'] = "IMPACTO:" + str(impacto)
    sheet['C5'] = "RIESGO:" + str(riesgoInforme)

    # DETALLES DE LA PRUEBA LABEL
    sheet.merge_cells('A6:C6')
    sheet.row_dimensions[6].height = 29
    sheet['A6'].font = Arial11Bold

    sheet['A6'].border = thin_border
    sheet['B6'].border = thin_border
    sheet['C6'].border = thin_border

    sheet['A6'].alignment = Alignment(horizontal="center", vertical='center', wrap_text=True)
    sheet['A6'] = "DETALLES DE LA PRUEBA"
    sheet['A6'].fill = greyFill

    # HOSTS LABEL
    hostHeight1 = 20 + 20 * (len(hosts) // 80)
    countNL = hosts.count("SALTOLINEA")
    hostHeight2 = (countNL + 1) * 20
    if hostHeight1 > hostHeight2:
        hostHeight = hostHeight1
    else:
        hostHeight = hostHeight2

    sheet['A7'].font = Arial11Bold
    sheet['B7'].font = Calibri10

    sheet['A7'].border = thin_border
    sheet['B7'].border = thin_border
    sheet['C7'].border = thin_border

    sheet['A7'].alignment = Alignment(horizontal="center", vertical='center', wrap_text=True)
    sheet['A7'] = "Hosts:"
    sheet.row_dimensions[7].height = hostHeight
    sheet.merge_cells('B7:C7')
    sheet['B7'].alignment = Alignment(horizontal="left", vertical='center', wrap_text=True)
    hosts = hosts.replace("SALTOLINEA", "\n")
    sheet['B7'] = hosts

    # DETALLES DE LA PRUEBA
    cellHeight = 20 + 20 * (len(detallesTest) // 80)
    sheet.row_dimensions[8].height = cellHeight

    sheet.merge_cells('A8:C8')
    sheet.row_dimensions[8].height = cellHeight
    sheet['A8'].font = Arial10

    sheet['A8'].border = thin_border
    sheet['B8'].border = thin_border
    sheet['C8'].border = thin_border

    sheet['A8'].alignment = Alignment(horizontal="justify", vertical='center', wrap_text=True)
    detallesTest = detallesTest.replace("SALTOLINEA", "\n")
    sheet['A8'] = detallesTest

    # CONTRAMEDIDAS - LABEL
    sheet.merge_cells('A9:C9')
    sheet.row_dimensions[9].height = 29
    sheet['A9'].font = Arial11Bold

    sheet['A9'].border = thin_border
    sheet['B9'].border = thin_border
    sheet['C9'].border = thin_border

    sheet['A9'].alignment = Alignment(horizontal="center", vertical='center', wrap_text=True)
    sheet['A9'] = "CONTRAMEDIDAS"
    sheet['A9'].fill = greyFill

    # CONTRAMEDIDAS
    sheet.merge_cells('A10:C10')
    cellHeight = 20 + 20 * (len(recomendacion) // 80)


    sheet.row_dimensions[10].height = cellHeight
    sheet['A10'].font = Arial10

    sheet['A10'].border = thin_border
    sheet['B10'].border = thin_border
    sheet['C10'].border = thin_border

    sheet['A10'].alignment = Alignment(horizontal="justify", vertical='center', wrap_text=True)
    recomendacion = recomendacion.replace("SALTOLINEA", "\n")
    sheet['A10'] = recomendacion

    # REFERENCIAS - LABEL
    sheet.merge_cells('A11:C11')
    sheet.row_dimensions[11].height = 29
    sheet['A11'].font = Arial11Bold

    sheet['A11'].border = thin_border
    sheet['B11'].border = thin_border
    sheet['C11'].border = thin_border

    sheet['A11'].alignment = Alignment(horizontal="center", vertical='center', wrap_text=True)
    sheet['A11'] = "REFERENCIAS"
    sheet['A11'].fill = greyFill

    # REFERENCIAS
    sheet.merge_cells('A12:C12')
    cellHeight = 20 + 20 * (len(referencias) // 80)
    sheet.row_dimensions[12].height = cellHeight
    sheet['A12'].font = Arial10

    sheet['A12'].border = thin_border
    sheet['B12'].border = thin_border
    sheet['C12'].border = thin_border

    sheet['A12'].alignment = Alignment(horizontal="justify", vertical='center', wrap_text=True)
    referencias = referencias.replace("SALTOLINEA", "\n")
    referencias = referencias.replace("TAB", "\t")
    sheet['A12'] = referencias    
wbISB.save('informeISB.xlsx')
